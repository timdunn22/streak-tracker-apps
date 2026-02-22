#!/usr/bin/env python3
"""
Comprehensive test suite for subscription-tracker.xlsx
Tests all 13 categories: sheets, formulas, validations, conditional formatting,
styling, frozen panes, column widths, number formats, row capacity,
cancellation log, renewal calendar, annual summary, auto-filter.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import sys

FILE_PATH = "/Users/timdunn/mobile_app_ideas/etsy-templates/subscription-tracker.xlsx"

passes = 0
fails = 0
fail_details = []

def check(test_name, condition, detail=""):
    global passes, fails
    if condition:
        passes += 1
        print(f"  PASS: {test_name}")
    else:
        fails += 1
        msg = f"  FAIL: {test_name}"
        if detail:
            msg += f" -- {detail}"
        print(msg)
        fail_details.append(f"{test_name}: {detail}")

def get_color_rgb(color_obj):
    """Extract hex RGB from openpyxl Color object."""
    if color_obj is None:
        return None
    if hasattr(color_obj, 'rgb') and color_obj.rgb and color_obj.type == 'rgb':
        rgb = color_obj.rgb
        if isinstance(rgb, str):
            return rgb[-6:].upper()
    return None

def parse_dv_list(formula1):
    """Parse data validation list formula1, stripping openpyxl's wrapping quotes."""
    cleaned = formula1.strip('"')
    return [c.strip() for c in cleaned.split(',')]


# ============================================================
print("=" * 70)
print("SUBSCRIPTION TRACKER -- COMPREHENSIVE TEST SUITE")
print("=" * 70)

wb = openpyxl.load_workbook(FILE_PATH)

# ============================================================
# 1. SHEET EXISTENCE
# ============================================================
print("\n--- 1. SHEET EXISTENCE ---")
expected_sheets = [
    "Instructions", "Dashboard", "All Subscriptions",
    "Renewal Calendar", "Annual Summary", "Cancellation Log"
]
for name in expected_sheets:
    check(f"Sheet '{name}' exists", name in wb.sheetnames, f"Sheets found: {wb.sheetnames}")
check("Exactly 6 sheets", len(wb.sheetnames) == 6, f"Found {len(wb.sheetnames)}: {wb.sheetnames}")


# ============================================================
# 2. FORMULAS (15+ examples)
# ============================================================
print("\n--- 2. FORMULAS (15+ examples) ---")

ws_d = wb['Dashboard']
ws_s = wb['All Subscriptions']
ws_cl = wb['Cancellation Log']
ws_rc = wb['Renewal Calendar']
ws_as = wb['Annual Summary']

# Dashboard KPI formulas
c5 = ws_d['C5'].value
check("Dashboard C5: Total Monthly Cost = SUMPRODUCT (non-canceled * monthly cost)",
      c5 and 'SUMPRODUCT' in str(c5) and 'Canceled' in str(c5) and 'C4:C103' in str(c5),
      f"Got: {c5}")
print(f"    >>> Formula: {c5}")

c6 = ws_d['C6'].value
check("Dashboard C6: Total Annual Cost = SUMPRODUCT (non-canceled * annual cost)",
      c6 and 'SUMPRODUCT' in str(c6) and 'D4:D103' in str(c6),
      f"Got: {c6}")
print(f"    >>> Formula: {c6}")

c7 = ws_d['C7'].value
check("Dashboard C7: Average Monthly Cost = IFERROR(SUMPRODUCT/COUNTIF)",
      c7 and 'IFERROR' in str(c7) and 'SUMPRODUCT' in str(c7) and 'COUNTIF' in str(c7),
      f"Got: {c7}")
print(f"    >>> Formula: {c7}")

c8 = ws_d['C8'].value
check("Dashboard C8: Most Expensive Monthly = IFERROR(MAX(...))",
      c8 and 'MAX' in str(c8) and 'C4:C103' in str(c8),
      f"Got: {c8}")
print(f"    >>> Formula: {c8}")

# Dashboard COUNTIF status counts
f5 = ws_d['F5'].value
check("Dashboard F5: Active count = COUNTIF(...,'Active')",
      f5 and 'COUNTIF' in str(f5) and 'Active' in str(f5),
      f"Got: {f5}")
print(f"    >>> Formula: {f5}")

f6 = ws_d['F6'].value
check("Dashboard F6: Paused count = COUNTIF(...,'Paused')",
      f6 and 'COUNTIF' in str(f6) and 'Paused' in str(f6),
      f"Got: {f6}")
print(f"    >>> Formula: {f6}")

f7 = ws_d['F7'].value
check("Dashboard F7: Consider Canceling count = COUNTIF",
      f7 and 'COUNTIF' in str(f7) and 'Consider Canceling' in str(f7),
      f"Got: {f7}")
print(f"    >>> Formula: {f7}")

f8 = ws_d['F8'].value
check("Dashboard F8: Canceled count = COUNTIF(...,'Canceled')",
      f8 and 'COUNTIF' in str(f8) and 'Canceled' in str(f8),
      f"Got: {f8}")
print(f"    >>> Formula: {f8}")

# Dashboard SUMPRODUCT category breakdown
c12 = ws_d['C12'].value
check("Dashboard C12: Entertainment count = COUNTIFS",
      c12 and 'COUNTIFS' in str(c12) and 'Entertainment' in str(c12),
      f"Got: {c12}")
print(f"    >>> Formula: {c12}")

e12 = ws_d['E12'].value
check("Dashboard E12: Entertainment monthly total = SUMPRODUCT",
      e12 and 'SUMPRODUCT' in str(e12) and 'Entertainment' in str(e12),
      f"Got: {e12}")
print(f"    >>> Formula: {e12}")

f12 = ws_d['F12'].value
check("Dashboard F12: Entertainment annual total = SUMPRODUCT",
      f12 and 'SUMPRODUCT' in str(f12) and 'Entertainment' in str(f12) and 'D4:D103' in str(f12),
      f"Got: {f12}")
print(f"    >>> Formula: {f12}")

# All Subscriptions Annual Cost formula
d4 = ws_s['D4'].value
check("All Subs D4: Annual Cost = IF(C4='','',...weekly/monthly/quarterly/annual)",
      d4 and 'IF' in str(d4) and 'Annual' in str(d4) and 'C4' in str(d4),
      f"Got: {d4}")
print(f"    >>> Formula: {d4}")

# Cancellation Log Annual Savings formula
d5_cl = ws_cl['D5'].value
check("Cancel Log D5: Annual Savings = IF(C5='',''C5*12)",
      d5_cl and 'C5*12' in str(d5_cl),
      f"Got: {d5_cl}")
print(f"    >>> Formula: {d5_cl}")

c3_cl = ws_cl['C3'].value
check("Cancel Log C3: Monthly savings total = SUM(C5:C104)",
      c3_cl and 'SUM' in str(c3_cl) and 'C5:C104' in str(c3_cl),
      f"Got: {c3_cl}")
print(f"    >>> Formula: {c3_cl}")

d3_cl = ws_cl['D3'].value
check("Cancel Log D3: Annual savings total = SUM(D5:D104)",
      d3_cl and 'SUM' in str(d3_cl) and 'D5:D104' in str(d3_cl),
      f"Got: {d3_cl}")
print(f"    >>> Formula: {d3_cl}")

# Annual Summary formulas
o6 = ws_as['O6'].value
check("Annual Summary O6: Year total = SUM(C6:N6)",
      o6 and 'SUM' in str(o6) and 'C6:N6' in str(o6),
      f"Got: {o6}")
print(f"    >>> Formula: {o6}")

c8_as = ws_as['C8'].value
check("Annual Summary C8: Difference = C6-C7",
      c8_as and 'C6-C7' in str(c8_as),
      f"Got: {c8_as}")
print(f"    >>> Formula: {c8_as}")

o12 = ws_as['O12'].value
check("Annual Summary O12: Entertainment category total = SUM(C12:N12)",
      o12 and 'SUM' in str(o12) and 'C12:N12' in str(o12),
      f"Got: {o12}")
print(f"    >>> Formula: {o12}")

c22_as = ws_as['C22'].value
check("Annual Summary C22: TOTAL column = SUM(C12:C21)",
      c22_as and 'SUM' in str(c22_as) and 'C12:C21' in str(c22_as),
      f"Got: {c22_as}")
print(f"    >>> Formula: {c22_as}")

c25 = ws_as['C25'].value
check("Annual Summary C25: Monthly savings from Cancel Log",
      c25 and 'Cancellation Log' in str(c25),
      f"Got: {c25}")
print(f"    >>> Formula: {c25}")

b6_rc = ws_rc['B6'].value
check("Renewal Calendar B6: IFERROR(INDEX(...SMALL(IF(MONTH=1...))))",
      b6_rc and 'IFERROR' in str(b6_rc) and 'INDEX' in str(b6_rc) and 'SMALL' in str(b6_rc),
      f"Got: {str(b6_rc)[:80]}")
print(f"    >>> Formula: {str(b6_rc)[:100]}")

# Count formulas printed
formula_count = 0
for cell_ref in ['C5','C6','C7','C8','F5','F6','F7','F8','C12','E12','F12',
                 'C13','E13','F13','C14','E14','F14']:
    v = ws_d[cell_ref].value
    if v and str(v).startswith('='):
        formula_count += 1
check(f"Dashboard has 15+ formula cells in sample (found {formula_count})",
      formula_count >= 15, f"Found {formula_count}")


# ============================================================
# 3. DATA VALIDATION (Dropdowns)
# ============================================================
print("\n--- 3. DATA VALIDATION (Dropdowns) ---")

dv_map = {}
for dv in ws_s.data_validations.dataValidation:
    sqref = str(dv.sqref)
    dv_map[sqref] = dv

# Category dropdown (B4:B103)
check("Category dropdown exists on B4:B103", 'B4:B103' in dv_map)
if 'B4:B103' in dv_map:
    actual_cats = parse_dv_list(dv_map['B4:B103'].formula1)
    expected_cats = ["Entertainment", "Software", "Health & Fitness", "News & Media",
                     "Food & Delivery", "Cloud Storage", "Music", "Gaming", "Productivity", "Other"]
    check("Category dropdown has 10 options", len(actual_cats) == 10, f"Got {len(actual_cats)}: {actual_cats}")
    for cat in expected_cats:
        check(f"  Category includes '{cat}'", cat in actual_cats, f"Actual: {actual_cats}")

# Billing Cycle dropdown (E4:E103)
check("Billing Cycle dropdown exists on E4:E103", 'E4:E103' in dv_map)
if 'E4:E103' in dv_map:
    bc_vals = parse_dv_list(dv_map['E4:E103'].formula1)
    expected_bc = ["Monthly", "Annual", "Quarterly", "Weekly"]
    check("Billing Cycle has 4 options", len(bc_vals) == 4, f"Got: {bc_vals}")
    for bc in expected_bc:
        check(f"  Billing Cycle includes '{bc}'", bc in bc_vals)

# Auto-Renew dropdown (G4:G103)
check("Auto-Renew dropdown exists on G4:G103", 'G4:G103' in dv_map)
if 'G4:G103' in dv_map:
    ar_vals = parse_dv_list(dv_map['G4:G103'].formula1)
    check("Auto-Renew has 2 options (Yes, No)",
          set(ar_vals) == {"Yes", "No"}, f"Got: {ar_vals}")

# Payment Method dropdown (H4:H103)
check("Payment Method dropdown exists on H4:H103", 'H4:H103' in dv_map)
if 'H4:H103' in dv_map:
    pm_vals = parse_dv_list(dv_map['H4:H103'].formula1)
    expected_pm = ["Credit Card", "Debit", "PayPal", "Apple Pay", "Google Pay", "Bank Transfer"]
    check("Payment Method has 6 options", len(pm_vals) == 6, f"Got: {pm_vals}")
    for pm in expected_pm:
        check(f"  Payment Method includes '{pm}'", pm in pm_vals)

# Status dropdown (I4:I103)
check("Status dropdown exists on I4:I103", 'I4:I103' in dv_map)
if 'I4:I103' in dv_map:
    st_vals = parse_dv_list(dv_map['I4:I103'].formula1)
    expected_st = ["Active", "Paused", "Consider Canceling", "Canceled"]
    check("Status has 4 options", len(st_vals) == 4, f"Got: {st_vals}")
    for st in expected_st:
        check(f"  Status includes '{st}'", st in st_vals)

# Cancellation Log: Would Re-subscribe
cl_dv_map = {}
for dv in ws_cl.data_validations.dataValidation:
    cl_dv_map[str(dv.sqref)] = dv

check("Cancel Log 'Would Re-subscribe?' dropdown on F5:F104", 'F5:F104' in cl_dv_map)
if 'F5:F104' in cl_dv_map:
    rs_vals = parse_dv_list(cl_dv_map['F5:F104'].formula1)
    check("  Would Re-subscribe has Yes/No/Maybe",
          set(rs_vals) == {"Yes", "No", "Maybe"}, f"Got: {rs_vals}")


# ============================================================
# 4. CONDITIONAL FORMATTING
# ============================================================
print("\n--- 4. CONDITIONAL FORMATTING ---")

status_cf_found = False
status_rules = {}
canceled_row_cf_count = 0

for cf in ws_s.conditional_formatting:
    for rule in cf.rules:
        if rule.type == 'cellIs' and rule.formula:
            formula_val = rule.formula[0].strip('"')
            if formula_val in ['Active', 'Paused', 'Consider Canceling', 'Canceled']:
                status_cf_found = True
                fg = get_color_rgb(rule.dxf.fill.fgColor) if rule.dxf and rule.dxf.fill else None
                font_c = get_color_rgb(rule.dxf.font.color) if rule.dxf and rule.dxf.font else None
                status_rules[formula_val] = {'fill': fg, 'font': font_c}

        if rule.type == 'expression' and rule.formula and 'Canceled' in str(rule.formula):
            canceled_row_cf_count += 1

check("Status column (I4:I103) has conditional formatting", status_cf_found)

if 'Active' in status_rules:
    r = status_rules['Active']
    check("Active status: green fill (#D1FAE5)", r['fill'] == 'D1FAE5', f"Got: {r['fill']}")
    check("Active status: green font (#059669)", r['font'] == '059669', f"Got: {r['font']}")

if 'Paused' in status_rules:
    r = status_rules['Paused']
    check("Paused status: yellow fill (#FEF3C7)", r['fill'] == 'FEF3C7', f"Got: {r['fill']}")
    check("Paused status: amber font (#D97706)", r['font'] == 'D97706', f"Got: {r['font']}")

if 'Consider Canceling' in status_rules:
    r = status_rules['Consider Canceling']
    check("Consider Canceling: orange fill (#FFEDD5)", r['fill'] == 'FFEDD5', f"Got: {r['fill']}")
    check("Consider Canceling: orange font (#EA580C)", r['font'] == 'EA580C', f"Got: {r['font']}")

if 'Canceled' in status_rules:
    r = status_rules['Canceled']
    check("Canceled status: gray fill (#F3F4F6)", r['fill'] == 'F3F4F6', f"Got: {r['fill']}")
    check("Canceled status: gray font (#9CA3AF)", r['font'] == '9CA3AF', f"Got: {r['font']}")

check("Canceled row CF applied to all data columns (A-H, J) = 9+ ranges",
      canceled_row_cf_count >= 9, f"Found {canceled_row_cf_count}")

# Category >$50/month highlighting on Dashboard
dash_cf_found = False
for cf in ws_d.conditional_formatting:
    for rule in cf.rules:
        if rule.type == 'cellIs' and rule.operator == 'greaterThan':
            if rule.formula and '50' in str(rule.formula):
                dash_cf_found = True
                fg = get_color_rgb(rule.dxf.fill.fgColor) if rule.dxf and rule.dxf.fill else None
                check("Categories >$50/mo highlighted on Dashboard (E12:E21)", True)
                check("  Highlight color is yellow/amber (#FDE68A)", fg == 'FDE68A', f"Got: {fg}")

check("Dashboard has >$50/month category highlighting", dash_cf_found)


# ============================================================
# 5. STYLING
# ============================================================
print("\n--- 5. STYLING ---")

for sheet_name in ["Dashboard", "All Subscriptions", "Renewal Calendar",
                   "Annual Summary", "Cancellation Log"]:
    ws = wb[sheet_name]
    c = ws['A1']
    fg = get_color_rgb(c.fill.fgColor) if c.fill and c.fill.fill_type == 'solid' else None
    check(f"{sheet_name} A1 has deep purple header fill (#4C1D95)",
          fg == '4C1D95', f"Got: {fg}")

for sheet_name in ["Dashboard", "All Subscriptions"]:
    ws = wb[sheet_name]
    c = ws['A1']
    font_c = get_color_rgb(c.font.color) if c.font and c.font.color else None
    check(f"{sheet_name} A1 header font is white (#FFFFFF)",
          font_c == 'FFFFFF', f"Got: {font_c}")

header_fill = get_color_rgb(ws_s['A3'].fill.fgColor) if ws_s['A3'].fill.fill_type == 'solid' else None
check("All Subs header row 3 fill is deep purple (#4C1D95)",
      header_fill == '4C1D95', f"Got: {header_fill}")
check("All Subs header row 3 font is white",
      get_color_rgb(ws_s['A3'].font.color) == 'FFFFFF')
check("All Subs header row 3 font is bold", ws_s['A3'].font.bold == True)

# Alternating rows
row4_fill = get_color_rgb(ws_s.cell(row=4, column=1).fill.fgColor)
row5_fill = get_color_rgb(ws_s.cell(row=5, column=1).fill.fgColor)
check("All Subs row 4 (even) has light purple fill (#F5F3FF)",
      row4_fill == 'F5F3FF', f"Got: {row4_fill}")
check("All Subs row 5 (odd) has white fill (#FFFFFF)",
      row5_fill == 'FFFFFF', f"Got: {row5_fill}")

row6_fill = get_color_rgb(ws_s.cell(row=6, column=1).fill.fgColor)
row7_fill = get_color_rgb(ws_s.cell(row=7, column=1).fill.fgColor)
check("All Subs row 6 = purple, row 7 = white (alternating continues)",
      row6_fill == 'F5F3FF' and row7_fill == 'FFFFFF',
      f"Row 6={row6_fill}, Row 7={row7_fill}")

row102_fill = get_color_rgb(ws_s.cell(row=102, column=1).fill.fgColor)
row103_fill = get_color_rgb(ws_s.cell(row=103, column=1).fill.fgColor)
check("Alternating rows extend to row 102-103",
      row102_fill in ('F5F3FF', 'FFFFFF') and row103_fill in ('F5F3FF', 'FFFFFF')
      and row102_fill != row103_fill,
      f"Row 102={row102_fill}, Row 103={row103_fill}")


# ============================================================
# 6. FROZEN PANES
# ============================================================
print("\n--- 6. FROZEN PANES ---")

freeze_expectations = {
    "Instructions": "A3",
    "Dashboard": "A3",
    "All Subscriptions": "A4",
    "Renewal Calendar": "A3",
    "Annual Summary": "B3",
    "Cancellation Log": "A5",
}
for sheet_name, expected_freeze in freeze_expectations.items():
    ws = wb[sheet_name]
    check(f"{sheet_name} frozen at {expected_freeze}",
          ws.freeze_panes == expected_freeze, f"Got: {ws.freeze_panes}")


# ============================================================
# 7. COLUMN WIDTHS
# ============================================================
print("\n--- 7. COLUMN WIDTHS ---")

col_widths = {}
for col_letter in ['A','B','C','D','E','F','G','H','I','J']:
    col_widths[col_letter] = ws_s.column_dimensions[col_letter].width

check("All Subs col A (Service Name) width >= 20", col_widths['A'] >= 20, f"Got: {col_widths['A']}")
check("All Subs col B (Category) width >= 15", col_widths['B'] >= 15, f"Got: {col_widths['B']}")
check("All Subs col C (Monthly Cost) width >= 12", col_widths['C'] >= 12, f"Got: {col_widths['C']}")
check("All Subs col I (Status) width >= 15", col_widths['I'] >= 15, f"Got: {col_widths['I']}")
check("All Subs col J (Notes) width >= 20", col_widths['J'] >= 20, f"Got: {col_widths['J']}")
check("All Subs: all columns width >= 10",
      all(w >= 10 for w in col_widths.values()), f"Widths: {col_widths}")

for col_letter in ['A','B','C','D','E','F']:
    w = ws_cl.column_dimensions[col_letter].width
    check(f"Cancel Log col {col_letter} width >= 12", w >= 12, f"Got: {w}")


# ============================================================
# 8. NUMBER FORMATTING
# ============================================================
print("\n--- 8. NUMBER FORMATTING ---")

currency_fmt = '"$"#,##0.00'
date_fmt = 'MM/DD/YYYY'

check("All Subs C4 (Monthly Cost) = currency", ws_s['C4'].number_format == currency_fmt, f"Got: {ws_s['C4'].number_format}")
check("All Subs D4 (Annual Cost) = currency", ws_s['D4'].number_format == currency_fmt, f"Got: {ws_s['D4'].number_format}")
check("All Subs F4 (Renewal Date) = MM/DD/YYYY", ws_s['F4'].number_format == date_fmt, f"Got: {ws_s['F4'].number_format}")
check("Currency format extends to row 50", ws_s.cell(row=50, column=3).number_format == currency_fmt)
check("Currency format extends to row 103", ws_s.cell(row=103, column=3).number_format == currency_fmt)
check("Date format extends to row 50", ws_s.cell(row=50, column=6).number_format == date_fmt)
check("Cancel Log C5 = currency", ws_cl['C5'].number_format == currency_fmt)
check("Cancel Log D5 = currency", ws_cl['D5'].number_format == currency_fmt)
check("Cancel Log B5 = date (MM/DD/YYYY)", ws_cl['B5'].number_format == date_fmt)
check("Dashboard C5 (Total Monthly) = currency", ws_d['C5'].number_format == currency_fmt)
check("Dashboard C6 (Total Annual) = currency", ws_d['C6'].number_format == currency_fmt)
check("Annual Summary C6 = currency", ws_as['C6'].number_format == currency_fmt)


# ============================================================
# 9. ROW CAPACITY
# ============================================================
print("\n--- 9. ROW CAPACITY ---")

formula_rows = sum(1 for r in range(4, 104) if ws_s.cell(row=r, column=4).value and str(ws_s.cell(row=r, column=4).value).startswith('='))
check("All Subs has 100 Annual Cost formulas (D4:D103)", formula_rows == 100, f"Found {formula_rows}")
check("All Subs row 103 has styling applied",
      get_color_rgb(ws_s.cell(row=103, column=1).fill.fgColor) in ('F5F3FF', 'FFFFFF'))
check("Category dropdown covers B4:B103 (100 rows)",
      any('B4:B103' in str(dv.sqref) for dv in ws_s.data_validations.dataValidation))


# ============================================================
# 10. CANCELLATION LOG
# ============================================================
print("\n--- 10. CANCELLATION LOG ---")

check("Cancel Log A4 = 'Service Name'", ws_cl['A4'].value == 'Service Name')
check("Cancel Log B4 = 'Date Canceled'", ws_cl['B4'].value == 'Date Canceled')
check("Cancel Log C4 = 'Monthly Savings'", ws_cl['C4'].value == 'Monthly Savings')
check("Cancel Log D4 = 'Annual Savings'", ws_cl['D4'].value == 'Annual Savings')
check("Cancel Log E4 = 'Reason'", ws_cl['E4'].value == 'Reason')
check("Cancel Log F4 = 'Would Re-subscribe?'", ws_cl['F4'].value == 'Would Re-subscribe?')

check("Cancel Log D5: Annual Savings = C5*12",
      ws_cl['D5'].value and 'C5*12' in str(ws_cl['D5'].value), f"Got: {ws_cl['D5'].value}")
check("Cancel Log D50 has Annual Savings formula",
      ws_cl.cell(row=50, column=4).value and '*12' in str(ws_cl.cell(row=50, column=4).value))
check("Cancel Log D104 has Annual Savings formula (last row)",
      ws_cl.cell(row=104, column=4).value and '*12' in str(ws_cl.cell(row=104, column=4).value))

cl_formula_count = sum(1 for r in range(5, 105)
                       if ws_cl.cell(row=r, column=4).value and '*12' in str(ws_cl.cell(row=r, column=4).value))
check("Cancel Log has 100 Annual Savings formulas (D5:D104)", cl_formula_count == 100, f"Found {cl_formula_count}")

check("Cancel Log C3 = SUM of monthly savings", ws_cl['C3'].value and 'SUM' in str(ws_cl['C3'].value))
check("Cancel Log D3 = SUM of annual savings", ws_cl['D3'].value and 'SUM' in str(ws_cl['D3'].value))
check("Cancel Log has 'Would Re-subscribe?' dropdown",
      any('F5:F104' in str(dv.sqref) for dv in ws_cl.data_validations.dataValidation))
check("Cancel Log has auto-filter enabled",
      ws_cl.auto_filter.ref is not None and ws_cl.auto_filter.ref != '')


# ============================================================
# 11. RENEWAL CALENDAR
# ============================================================
print("\n--- 11. RENEWAL CALENDAR ---")

expected_months = ['JANUARY','FEBRUARY','MARCH','APRIL','MAY','JUNE',
                   'JULY','AUGUST','SEPTEMBER','OCTOBER','NOVEMBER','DECEMBER']
month_rows = {}
for row in ws_rc.iter_rows(min_row=1, max_row=ws_rc.max_row, max_col=5, values_only=False):
    for c in row:
        if c.value in expected_months:
            month_rows[c.value] = c.row

for m in expected_months:
    check(f"Renewal Calendar has '{m}' section", m in month_rows)
check("All 12 monthly sections present", len(month_rows) == 12, f"Found {len(month_rows)}")

# January formulas
jan_row = month_rows.get('JANUARY')
if jan_row:
    fr = jan_row + 2
    b_val = ws_rc.cell(row=fr, column=2).value
    check(f"January has INDEX/SMALL formula at B{fr}",
          b_val and 'INDEX' in str(b_val) and 'MONTH' in str(b_val),
          f"Got: {str(b_val)[:60]}")

# December formulas
dec_row = month_rows.get('DECEMBER')
if dec_row:
    fr = dec_row + 2
    b_val = ws_rc.cell(row=fr, column=2).value
    check(f"December has INDEX/SMALL formula at B{fr}",
          b_val and 'INDEX' in str(b_val) and 'MONTH' in str(b_val),
          f"Got: {str(b_val)[:60]}")

# Headers
if jan_row:
    check("January has 'Service' header", ws_rc.cell(row=jan_row+1, column=2).value == 'Service')
    check("January has 'Amount' header", ws_rc.cell(row=jan_row+1, column=5).value == 'Amount')

# July references month 7
jul_row = month_rows.get('JULY')
if jul_row:
    b_val = str(ws_rc.cell(row=jul_row+2, column=2).value or '')
    check("July formulas reference MONTH(...)=7",
          '=7' in b_val or ',7,' in b_val or '7)' in b_val, f"Got: {b_val[:80]}")


# ============================================================
# 12. ANNUAL SUMMARY
# ============================================================
print("\n--- 12. ANNUAL SUMMARY ---")

check("Annual Summary has 'MONTH-BY-MONTH SPENDING' header",
      ws_as['B4'].value == 'MONTH-BY-MONTH SPENDING')

# Check month headers are properly aligned (C5:N5 after fix)
month_abbrevs = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
found_months_row5 = []
for col in range(1, 16):
    v = ws_as.cell(row=5, column=col).value
    if v in month_abbrevs:
        found_months_row5.append((col, v))

check("Annual Summary row 5 has all 12 month abbreviations",
      len(found_months_row5) == 12, f"Found {len(found_months_row5)}")

total_col_r5 = None
for col in range(1, 16):
    if ws_as.cell(row=5, column=col).value == 'TOTAL':
        total_col_r5 = col
check("Annual Summary row 5 has 'TOTAL' column", total_col_r5 is not None)

# Verify alignment: first month should be in col C (3)
if found_months_row5:
    first_month_col = found_months_row5[0][0]
    check("Month headers aligned with data columns (Jan starts at C5)",
          first_month_col == 3,
          f"First month at column {first_month_col} (expected 3/C)")

# TOTAL should be in col O (15)
check("TOTAL header in col O (15) to match SUM formulas",
      total_col_r5 == 15, f"TOTAL at col {total_col_r5}")

check("Annual Summary B6 = 'Current Year'", ws_as['B6'].value == 'Current Year')
check("Annual Summary B7 = 'Previous Year'", ws_as['B7'].value == 'Previous Year')
check("Annual Summary B8 = 'Difference (+/-)'", ws_as['B8'].value == 'Difference (+/-)')

diff_count = sum(1 for col in range(3, 15)
                 if ws_as.cell(row=8, column=col).value and '-' in str(ws_as.cell(row=8, column=col).value))
check("Annual Summary row 8 has 12 difference formulas", diff_count == 12, f"Found {diff_count}")

check("Annual Summary O6 = SUM for current year", ws_as['O6'].value and 'SUM' in str(ws_as['O6'].value))
check("Annual Summary O7 = SUM for previous year", ws_as['O7'].value and 'SUM' in str(ws_as['O7'].value))

check("Annual Summary B10 = 'CATEGORY-BY-CATEGORY...'",
      ws_as['B10'].value and 'CATEGORY' in str(ws_as['B10'].value))

# Category row 11 month headers aligned (should start at C11 after fix)
found_months_row11 = []
for col in range(1, 16):
    v = ws_as.cell(row=11, column=col).value
    if v in month_abbrevs:
        found_months_row11.append((col, v))
if found_months_row11:
    check("Category month headers aligned (Jan at C11)",
          found_months_row11[0][0] == 3,
          f"First month at column {found_months_row11[0][0]}")

categories = ["Entertainment", "Software", "Health & Fitness", "News & Media",
              "Food & Delivery", "Cloud Storage", "Music", "Gaming", "Productivity", "Other"]
for i, cat in enumerate(categories):
    row = 12 + i
    check(f"Annual Summary B{row} = '{cat}'", ws_as.cell(row=row, column=2).value == cat,
          f"Got: {ws_as.cell(row=row, column=2).value}")

check("Annual Summary B22 = 'TOTAL'", ws_as['B22'].value == 'TOTAL')

total_formulas = sum(1 for col in range(3, 16)
                     if ws_as.cell(row=22, column=col).value and 'SUM' in str(ws_as.cell(row=22, column=col).value))
check("Annual Summary row 22 has 13 SUM totals (C22:O22)", total_formulas == 13, f"Found {total_formulas}")

cat_totals = sum(1 for row in range(12, 22)
                 if ws_as.cell(row=row, column=15).value and 'SUM' in str(ws_as.cell(row=row, column=15).value))
check("10 category row totals in col O", cat_totals == 10, f"Found {cat_totals}")

check("Annual Summary B25 = 'Monthly Savings (from log)'",
      ws_as['B25'].value and 'Monthly Savings' in str(ws_as['B25'].value))
check("Annual Summary C25 refs Cancellation Log",
      ws_as['C25'].value and 'Cancellation Log' in str(ws_as['C25'].value))
check("Annual Summary C26 refs Cancellation Log (annual)",
      ws_as['C26'].value and 'Cancellation Log' in str(ws_as['C26'].value))


# ============================================================
# 13. AUTO-FILTER
# ============================================================
print("\n--- 13. AUTO-FILTER ---")

check("All Subscriptions auto-filter enabled",
      ws_s.auto_filter.ref is not None and ws_s.auto_filter.ref != '')
check("All Subscriptions auto-filter covers A3:J103",
      ws_s.auto_filter.ref == 'A3:J103', f"Got: {ws_s.auto_filter.ref}")
check("Cancellation Log auto-filter covers A4:F104",
      ws_cl.auto_filter.ref == 'A4:F104', f"Got: {ws_cl.auto_filter.ref}")


# ============================================================
# SUMMARY
# ============================================================
print("\n" + "=" * 70)
print(f"RESULTS: {passes} PASSED, {fails} FAILED out of {passes + fails} total checks")
print("=" * 70)

if fails > 0:
    print("\nFAILURES:")
    for fd in fail_details:
        print(f"  * {fd}")
    sys.exit(1)
else:
    print("\nAll checks passed!")
    sys.exit(0)
