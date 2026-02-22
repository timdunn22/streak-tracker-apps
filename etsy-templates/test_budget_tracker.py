#!/usr/bin/env python3
"""Comprehensive test suite for monthly-budget-tracker.xlsx"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import sys
import os

FILEPATH = "/Users/timdunn/mobile_app_ideas/etsy-templates/monthly-budget-tracker.xlsx"
passes = 0
fails = 0
issues = []

def log_pass(msg):
    global passes
    passes += 1
    print(f"  PASS: {msg}")

def log_fail(msg):
    global fails
    fails += 1
    issues.append(msg)
    print(f"  FAIL: {msg}")

wb = openpyxl.load_workbook(FILEPATH)
wb_data = openpyxl.load_workbook(FILEPATH, data_only=True)

print("=" * 80)
print("COMPREHENSIVE TEST: monthly-budget-tracker.xlsx")
print("=" * 80)

# ============================================================
# 1. SHEET EXISTENCE
# ============================================================
print("\n--- TEST 1: Sheet Existence ---")
expected_sheets = ["Instructions", "Dashboard", "Monthly Budget", "Transactions", "Annual Overview"]
actual_sheets = wb.sheetnames

for sheet_name in expected_sheets:
    if sheet_name in actual_sheets:
        log_pass(f"Sheet '{sheet_name}' exists")
    else:
        log_fail(f"Sheet '{sheet_name}' is MISSING")

extra_sheets = set(actual_sheets) - set(expected_sheets)
if extra_sheets:
    print(f"  INFO: Extra sheets found: {extra_sheets}")
else:
    log_pass("No unexpected extra sheets")

print(f"  INFO: Sheet order: {actual_sheets}")

# ============================================================
# 2. FORMULAS
# ============================================================
print("\n--- TEST 2: Formulas ---")

formula_checks = []

for sheet_name in ["Dashboard", "Monthly Budget", "Annual Overview", "Transactions"]:
    if sheet_name not in actual_sheets:
        continue
    ws = wb[sheet_name]
    print(f"  INFO: {sheet_name} dimensions: {ws.dimensions}")
    formula_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 520), max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formula_checks.append((sheet_name, cell.coordinate, cell.value[:80]))
                formula_count += 1
    if formula_count > 0:
        log_pass(f"{sheet_name} has {formula_count} formula cells")
    elif sheet_name == "Transactions":
        print(f"  INFO: Transactions has {formula_count} formula cells (may be expected)")
    else:
        log_fail(f"{sheet_name} has NO formula cells")

print(f"\n  Formula examples (showing up to 15 of {len(formula_checks)} total):")
for sheet, coord, formula in formula_checks[:15]:
    print(f"    [{sheet}] {coord}: {formula}")

if len(formula_checks) >= 10:
    log_pass(f"Found {len(formula_checks)} total formulas across sheets (>=10 required)")
else:
    log_fail(f"Found only {len(formula_checks)} total formulas (need at least 10)")

# ============================================================
# 3. DATA VALIDATION (Dropdowns)
# ============================================================
print("\n--- TEST 3: Data Validation (Dropdowns) ---")

if "Transactions" in actual_sheets:
    ws = wb["Transactions"]
    validations = ws.data_validations.dataValidation if ws.data_validations else []
    print(f"  INFO: Found {len(validations)} data validation rules on Transactions sheet")
    
    found_category_dv = False
    found_payment_dv = False
    
    for dv in validations:
        ranges_str = str(dv.sqref)
        print(f"    Validation on {ranges_str}: type={dv.type}, formula1={str(dv.formula1)[:120] if dv.formula1 else 'None'}")
        
        if dv.formula1:
            formula_lower = str(dv.formula1).lower()
            if any(cat in formula_lower for cat in ['housing', 'food', 'groceries', 'transport', 'utilities', 'entertainment', 'income', 'salary']):
                found_category_dv = True
                log_pass(f"Category dropdown found on {ranges_str}")
                print(f"      Options: {dv.formula1[:200]}")
            elif any(pay in formula_lower for pay in ['cash', 'credit', 'debit', 'bank', 'check', 'venmo', 'paypal', 'transfer']):
                found_payment_dv = True
                log_pass(f"Payment Method dropdown found on {ranges_str}")
                print(f"      Options: {dv.formula1[:200]}")
    
    if not found_category_dv:
        log_fail("No Category dropdown validation found on Transactions sheet")
    if not found_payment_dv:
        log_fail("No Payment Method dropdown validation found on Transactions sheet")
    
    if not validations:
        log_fail("NO data validations found on Transactions sheet at all")

# ============================================================
# 4. CONDITIONAL FORMATTING
# ============================================================
print("\n--- TEST 4: Conditional Formatting ---")

sheets_with_cf_expected = ["Dashboard", "Monthly Budget", "Transactions", "Annual Overview"]
for sheet_name in sheets_with_cf_expected:
    if sheet_name in actual_sheets:
        ws = wb[sheet_name]
        cf_rules = ws.conditional_formatting
        rule_count = len(list(cf_rules))
        if rule_count > 0:
            log_pass(f"'{sheet_name}' has {rule_count} conditional formatting rule(s)")
            for cf in cf_rules:
                for rule in cf.rules:
                    print(f"    Range: {cf.sqref}, Type: {rule.type}, Priority: {rule.priority}")
        else:
            log_fail(f"'{sheet_name}' has NO conditional formatting rules")

# ============================================================
# 5. SAMPLE DATA (16 transactions)
# ============================================================
print("\n--- TEST 5: Sample Transactions Data ---")

if "Transactions" in actual_sheets:
    ws_data = wb_data["Transactions"]
    
    header_row = None
    headers = {}
    for row_idx in range(1, 5):
        for col_idx in range(1, 15):
            val = ws_data.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip().lower() in ['date', 'category', 'description', 'amount', 'payment method', 'type', 'notes']:
                header_row = row_idx
                break
        if header_row:
            break
    
    if header_row:
        for col_idx in range(1, 15):
            val = ws_data.cell(row=header_row, column=col_idx).value
            if val:
                headers[str(val).strip()] = col_idx
        print(f"  INFO: Header row: {header_row}, Headers: {headers}")
        log_pass(f"Found header row at row {header_row}")
        
        data_rows = 0
        for row_idx in range(header_row + 1, header_row + 100):
            has_data = False
            for col_idx in range(1, min(len(headers) + 1, 10)):
                val = ws_data.cell(row=row_idx, column=col_idx).value
                if val is not None and str(val).strip() != "":
                    has_data = True
                    break
            if has_data:
                data_rows += 1
            else:
                break
        
        print(f"  INFO: Found {data_rows} consecutive data rows after header")
        if data_rows >= 16:
            log_pass(f"Found {data_rows} sample transactions (>=16 expected)")
        elif data_rows > 0:
            log_fail(f"Found only {data_rows} sample transactions (expected 16)")
        else:
            log_fail("No sample transaction data found")
        
        print("  Sample transactions:")
        for row_idx in range(header_row + 1, min(header_row + 6, header_row + data_rows + 1)):
            row_data = []
            for col_idx in range(1, min(len(headers) + 2, 8)):
                val = ws_data.cell(row=row_idx, column=col_idx).value
                row_data.append(str(val)[:20] if val else "")
            print(f"    Row {row_idx}: {' | '.join(row_data)}")
    else:
        log_fail("Could not find header row in Transactions sheet")

# ============================================================
# 6. STYLING
# ============================================================
print("\n--- TEST 6: Styling (Header fill=#1B2A4A navy, white font, Calibri) ---")

EXPECTED_FILL = "1B2A4A"

for sheet_name in expected_sheets:
    if sheet_name not in actual_sheets:
        continue
    ws = wb[sheet_name]
    
    found_styled_header = False
    for row_idx in range(1, 5):
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.fill and cell.fill.fgColor:
                fill_rgb = str(cell.fill.fgColor.rgb) if cell.fill.fgColor.rgb else ""
                fill_hex = fill_rgb[-6:].upper() if len(fill_rgb) >= 6 else fill_rgb.upper()
                
                if fill_hex == EXPECTED_FILL.upper():
                    found_styled_header = True
                    
                    font_color_rgb = ""
                    if cell.font and cell.font.color and cell.font.color.rgb:
                        try:
                            font_color_rgb = str(cell.font.color.rgb)[-6:].upper()
                        except:
                            font_color_rgb = "UNKNOWN"
                    
                    font_name = cell.font.name if cell.font else "Unknown"
                    
                    if font_color_rgb != "FFFFFF":
                        print(f"    [{sheet_name}] {cell.coordinate}: fill OK but font color is {font_color_rgb} (expected FFFFFF)")
                    if font_name and "Calibri" not in str(font_name):
                        print(f"    [{sheet_name}] {cell.coordinate}: font is '{font_name}' (expected Calibri)")
                    break
        if found_styled_header:
            break
    
    if found_styled_header:
        log_pass(f"'{sheet_name}' has navy (#1B2A4A) header styling")
    else:
        colors_found = set()
        for row_idx in range(1, 5):
            for col_idx in range(1, 10):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb)
                    if rgb != "00000000":
                        colors_found.add(rgb)
        log_fail(f"'{sheet_name}' missing navy header fill. Colors found: {colors_found if colors_found else 'none'}")

# Detailed style audit on Dashboard
print("  Detailed style audit on Dashboard row 1:")
if "Dashboard" in actual_sheets:
    ws = wb["Dashboard"]
    for col_idx in range(1, 4):
        cell = ws.cell(row=1, column=col_idx)
        fill_rgb = str(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb else "none"
        try:
            font_rgb = str(cell.font.color.rgb) if cell.font and cell.font.color and cell.font.color.rgb else "none"
        except:
            font_rgb = "theme/none"
        font_name = cell.font.name if cell.font else "none"
        font_bold = cell.font.bold if cell.font else False
        print(f"    {cell.coordinate}: val='{cell.value}' fill={fill_rgb} font_color={font_rgb} font={font_name} bold={font_bold}")

# ============================================================
# 7. FROZEN PANES
# ============================================================
print("\n--- TEST 7: Frozen Panes ---")

for sheet_name in expected_sheets:
    if sheet_name not in actual_sheets:
        continue
    ws = wb[sheet_name]
    freeze = ws.freeze_panes
    if freeze:
        log_pass(f"'{sheet_name}' has freeze panes at {freeze}")
    else:
        log_fail(f"'{sheet_name}' has NO freeze panes set")

# ============================================================
# 8. COLUMN WIDTHS
# ============================================================
print("\n--- TEST 8: Column Widths ---")

DEFAULT_WIDTH = 8.43

for sheet_name in expected_sheets:
    if sheet_name not in actual_sheets:
        continue
    ws = wb[sheet_name]
    custom_widths = 0
    width_details = []
    
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in ws.column_dimensions:
            w = ws.column_dimensions[col_letter].width
            if w and w != DEFAULT_WIDTH:
                custom_widths += 1
                width_details.append(f"{col_letter}={w:.1f}")
    
    if custom_widths > 0:
        log_pass(f"'{sheet_name}' has {custom_widths} custom column widths: {', '.join(width_details[:8])}")
    else:
        log_fail(f"'{sheet_name}' has ALL default column widths")

# ============================================================
# 9. NUMBER FORMATTING
# ============================================================
print("\n--- TEST 9: Number Formatting (currency) ---")

for sheet_name in ["Dashboard", "Monthly Budget", "Transactions", "Annual Overview"]:
    if sheet_name not in actual_sheets:
        continue
    ws = wb[sheet_name]
    currency_cells = 0
    formats_found = set()
    
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), max_col=ws.max_column):
        for cell in row:
            if cell.number_format and '$' in str(cell.number_format):
                currency_cells += 1
                formats_found.add(cell.number_format)
    
    if currency_cells > 0:
        log_pass(f"'{sheet_name}' has {currency_cells} cells with currency formatting")
        for fmt in formats_found:
            print(f"    Format: {fmt}")
    else:
        log_fail(f"'{sheet_name}' has NO currency-formatted cells")

# ============================================================
# 10. PRINT SETTINGS (Page Margins)
# ============================================================
print("\n--- TEST 10: Print Settings (Page Margins) ---")

for sheet_name in expected_sheets:
    if sheet_name not in actual_sheets:
        continue
    ws = wb[sheet_name]
    margins = ws.page_margins
    if margins:
        has_custom = (margins.left != 0.7 or margins.right != 0.7 or 
                     margins.top != 0.75 or margins.bottom != 0.75)
        margin_str = f"L={margins.left}, R={margins.right}, T={margins.top}, B={margins.bottom}"
        if has_custom:
            log_pass(f"'{sheet_name}' has custom margins: {margin_str}")
        else:
            log_pass(f"'{sheet_name}' has page margins set (default values): {margin_str}")
    else:
        log_fail(f"'{sheet_name}' has NO page margins")

# ============================================================
# 11. CELL MERGES (Instructions sheet)
# ============================================================
print("\n--- TEST 11: Cell Merges (Instructions sheet) ---")

if "Instructions" in actual_sheets:
    ws = wb["Instructions"]
    merged = list(ws.merged_cells.ranges)
    merge_count = len(merged)
    if merge_count > 0:
        log_pass(f"Instructions sheet has {merge_count} merged cell range(s)")
        for mr in merged:
            print(f"    Merged: {mr}")
    else:
        log_fail("Instructions sheet has NO merged cells")

for sheet_name in ["Dashboard", "Monthly Budget", "Annual Overview"]:
    if sheet_name in actual_sheets:
        ws = wb[sheet_name]
        merged = list(ws.merged_cells.ranges)
        if merged:
            print(f"  INFO: '{sheet_name}' has {len(merged)} merged ranges: {[str(m) for m in merged[:5]]}")

# ============================================================
# 12. ROW COUNT (Transactions 500+ formatted rows)
# ============================================================
print("\n--- TEST 12: Row Count (Transactions sheet 500+ formatted rows) ---")

if "Transactions" in actual_sheets:
    ws = wb["Transactions"]
    max_row = ws.max_row
    print(f"  INFO: Transactions max_row = {max_row}")
    
    # Count rows with any formatting or data
    formatted_rows = 0
    for row_idx in range(1, max_row + 1):
        has_format = False
        for col_idx in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if (cell.value is not None or
                (cell.fill and cell.fill.fgColor and str(cell.fill.fgColor.rgb) != "00000000") or
                (cell.number_format and cell.number_format != "General") or
                (cell.border and (cell.border.left.style or cell.border.right.style or 
                                  cell.border.top.style or cell.border.bottom.style))):
                has_format = True
                break
        if has_format:
            formatted_rows += 1
    
    print(f"  INFO: Transactions formatted rows = {formatted_rows}")
    
    if formatted_rows >= 500:
        log_pass(f"Transactions has {formatted_rows} formatted rows (>=500)")
    else:
        log_fail(f"Transactions has only {formatted_rows} formatted rows (need 500+)")

    # Also check if data validations cover 500 rows
    validations = ws.data_validations.dataValidation if ws.data_validations else []
    for dv in validations:
        print(f"  INFO: Data validation range: {dv.sqref}")

# ============================================================
# SUMMARY
# ============================================================
print("\n" + "=" * 80)
print(f"TEST SUMMARY: {passes} PASSED, {fails} FAILED")
print("=" * 80)

if issues:
    print("\nFailed items that need fixing:")
    for i, issue in enumerate(issues, 1):
        print(f"  {i}. {issue}")
else:
    print("\nAll tests passed!")

# Save issue list for potential fix script
with open("/Users/timdunn/mobile_app_ideas/etsy-templates/test_results.txt", "w") as f:
    f.write(f"passes={passes}\n")
    f.write(f"fails={fails}\n")
    for issue in issues:
        f.write(f"ISSUE: {issue}\n")

wb.close()
wb_data.close()
print(f"\nTotal formulas cataloged: {len(formula_checks)}")
