#!/usr/bin/env python3
"""
Subscription Tracker – Etsy-ready Excel template
Generates a professional, formula-rich .xlsx with 6 sheets.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from copy import copy

# ── colour palette ──────────────────────────────────────────────
DEEP_PURPLE   = "4C1D95"
MID_PURPLE    = "7C3AED"
LIGHT_PURPLE  = "F5F3FF"
VERY_LIGHT_P  = "FAF5FF"
EMERALD       = "059669"
LIGHT_GREEN   = "D1FAE5"
AMBER         = "D97706"
LIGHT_AMBER   = "FEF3C7"
ORANGE        = "EA580C"
LIGHT_ORANGE  = "FFEDD5"
RED           = "DC2626"
LIGHT_RED     = "FEE2E2"
GRAY          = "9CA3AF"
LIGHT_GRAY    = "F3F4F6"
WHITE         = "FFFFFF"
BLACK         = "000000"
DARK_TEXT      = "1F2937"

# Category colours (fill, font pairs)
CAT_COLORS = {
    "Entertainment":     ("EDE9FE", "5B21B6"),
    "Software":          ("DBEAFE", "1D4ED8"),
    "Health & Fitness":  ("D1FAE5", "047857"),
    "News & Media":      ("FEF3C7", "92400E"),
    "Food & Delivery":   ("FFEDD5", "9A3412"),
    "Cloud Storage":     ("E0E7FF", "3730A3"),
    "Music":             ("FCE7F3", "9D174D"),
    "Gaming":            ("CFFAFE", "0E7490"),
    "Productivity":      ("F3E8FF", "7E22CE"),
    "Other":             ("F3F4F6", "374151"),
}

CATEGORIES = list(CAT_COLORS.keys())

# ── reusable helpers ────────────────────────────────────────────
thin_border  = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

medium_border = Border(
    left=Side(style="medium", color=DEEP_PURPLE),
    right=Side(style="medium", color=DEEP_PURPLE),
    top=Side(style="medium", color=DEEP_PURPLE),
    bottom=Side(style="medium", color=DEEP_PURPLE),
)

header_font    = Font(name="Aptos", bold=True, color=WHITE, size=12)
header_fill    = PatternFill("solid", fgColor=DEEP_PURPLE)
sub_header_font = Font(name="Aptos", bold=True, color=DEEP_PURPLE, size=11)
sub_header_fill = PatternFill("solid", fgColor="EDE9FE")
body_font      = Font(name="Aptos", color=DARK_TEXT, size=11)
currency_fmt   = '"$"#,##0.00'
date_fmt       = 'MM/DD/YYYY'
center         = Alignment(horizontal="center", vertical="center")
left_center    = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_center   = Alignment(horizontal="right", vertical="center")


def style_header_row(ws, row, max_col, font=None, fill=None, height=32):
    """Apply header styling across a row."""
    f = font or header_font
    fl = fill or header_fill
    ws.row_dimensions[row].height = height
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = f
        c.fill = fl
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = medium_border


def style_data_cell(cell, row_idx, is_currency=False, is_date=False, is_center=False):
    """Style a data cell with alternating rows."""
    cell.font = body_font
    cell.border = thin_border
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=LIGHT_PURPLE)
    else:
        cell.fill = PatternFill("solid", fgColor=WHITE)
    if is_currency:
        cell.number_format = currency_fmt
        cell.alignment = right_center
    elif is_date:
        cell.number_format = date_fmt
        cell.alignment = center
    elif is_center:
        cell.alignment = center
    else:
        cell.alignment = left_center


def write_section_header(ws, row, col_start, col_end, text, merge=True):
    """Write a purple section header across merged cells."""
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = Font(name="Aptos", bold=True, color=WHITE, size=13)
    cell.fill = PatternFill("solid", fgColor=MID_PURPLE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = medium_border
    if merge and col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)
        for c in range(col_start + 1, col_end + 1):
            ws.cell(row=row, column=c).border = medium_border
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=MID_PURPLE)


def write_kpi_box(ws, row, col, label, formula, label_width=28, val_width=18):
    """Write a label+value KPI pair."""
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(name="Aptos", bold=True, color=DEEP_PURPLE, size=11)
    lc.fill = PatternFill("solid", fgColor="EDE9FE")
    lc.alignment = left_center
    lc.border = medium_border

    vc = ws.cell(row=row, column=col + 1)
    vc.value = formula
    vc.font = Font(name="Aptos", bold=True, color=DARK_TEXT, size=13)
    vc.fill = PatternFill("solid", fgColor=WHITE)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = medium_border
    return vc


# ═══════════════════════════════════════════════════════════════
wb = Workbook()

# ── SHEET 1: Instructions ─────────────────────────────────────
ws_instr = wb.active
ws_instr.title = "Instructions"
ws_instr.sheet_properties.tabColor = DEEP_PURPLE

# Page setup
ws_instr.page_margins = openpyxl.worksheet.page.PageMargins(
    left=0.5, right=0.5, top=0.5, bottom=0.5
)

# Column widths
ws_instr.column_dimensions["A"].width = 3
ws_instr.column_dimensions["B"].width = 90

# Title banner
ws_instr.merge_cells("A1:B1")
ws_instr.row_dimensions[1].height = 55
c = ws_instr["A1"]
c.value = "SUBSCRIPTION TRACKER"
c.font = Font(name="Aptos", bold=True, color=WHITE, size=24)
c.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_instr["B1"].fill = PatternFill("solid", fgColor=DEEP_PURPLE)
ws_instr["B1"].border = medium_border
ws_instr["A1"].border = medium_border

ws_instr.merge_cells("A2:B2")
ws_instr.row_dimensions[2].height = 30
c2 = ws_instr["A2"]
c2.value = "Take control of your recurring expenses. Find hidden costs. Save hundreds per year."
c2.font = Font(name="Aptos", italic=True, color=MID_PURPLE, size=12)
c2.fill = PatternFill("solid", fgColor="EDE9FE")
c2.alignment = Alignment(horizontal="center", vertical="center")

row = 4
sections = [
    ("WELCOME", [
        "Thank you for purchasing this Subscription Tracker! This workbook helps you catalog every",
        "recurring charge, visualize where your money goes, and identify subscriptions to cut.",
        "The average person has 12 active subscriptions and wastes $30-100/month on forgotten ones.",
    ]),
    ("HOW TO USE EACH SHEET", [
        "1. ALL SUBSCRIPTIONS  --  Enter every subscription you have. Dropdowns help you fill in",
        "   categories, billing cycles, and status. Formulas auto-calculate annual costs.",
        "",
        "2. DASHBOARD  --  See totals, averages, and category breakdowns update automatically.",
        "   No editing needed here; everything pulls from your subscription list.",
        "",
        "3. RENEWAL CALENDAR  --  View which subscriptions renew each month so you can plan ahead",
        "   and avoid surprise charges.",
        "",
        "4. ANNUAL SUMMARY  --  Track month-by-month spending and compare year-over-year.",
        "",
        "5. CANCELLATION LOG  --  Record every subscription you cancel and watch your savings grow!",
    ]),
    ("CATEGORY COLOR LEGEND", []),
    ("STATUS COLORS", [
        "Active  =  Green background (you are currently paying for this)",
        "Paused  =  Yellow background (temporarily on hold)",
        "Consider Canceling  =  Orange background (review this one!)",
        "Canceled  =  Gray background with strikethrough (done!)",
    ]),
    ("TIPS FOR CUTTING SUBSCRIPTIONS", [
        "1. Audit ruthlessly: If you haven't used it in 30 days, mark it \"Consider Canceling\".",
        "2. Check for duplicates: Do you have both Spotify and Apple Music?",
        "3. Downgrade first: Switch to a free tier before canceling entirely.",
        "4. Share family plans: Split Netflix, Spotify, or iCloud with family members.",
        "5. Negotiate: Call and ask for a retention discount -- it works more often than you think.",
        "6. Use annual billing: Many services offer 15-30% off when you pay yearly.",
        "7. Set calendar reminders: Review this tracker on the 1st of every month.",
        "8. Watch for price increases: Companies quietly raise prices; check statements quarterly.",
        "9. Free trial trap: Set a reminder to cancel 1 day before any free trial ends.",
        "10. The $5 rule: Small charges add up. Five $5 subscriptions = $300/year.",
    ]),
]

for title, lines in sections:
    ws_instr.merge_cells(f"A{row}:B{row}")
    ws_instr.row_dimensions[row].height = 30
    hc = ws_instr.cell(row=row, column=1, value=title)
    hc.font = Font(name="Aptos", bold=True, color=WHITE, size=13)
    hc.fill = PatternFill("solid", fgColor=MID_PURPLE)
    hc.alignment = Alignment(horizontal="left", vertical="center")
    hc.border = medium_border
    ws_instr.cell(row=row, column=2).fill = PatternFill("solid", fgColor=MID_PURPLE)
    ws_instr.cell(row=row, column=2).border = medium_border
    row += 1

    if title == "CATEGORY COLOR LEGEND":
        for cat_name, (bg, fg) in CAT_COLORS.items():
            ws_instr.merge_cells(f"A{row}:B{row}")
            ws_instr.row_dimensions[row].height = 24
            cc = ws_instr.cell(row=row, column=1, value=f"  {cat_name}")
            cc.font = Font(name="Aptos", bold=True, color=fg, size=11)
            cc.fill = PatternFill("solid", fgColor=bg)
            cc.alignment = left_center
            cc.border = thin_border
            ws_instr.cell(row=row, column=2).fill = PatternFill("solid", fgColor=bg)
            ws_instr.cell(row=row, column=2).border = thin_border
            row += 1
    else:
        for line in lines:
            ws_instr.merge_cells(f"A{row}:B{row}")
            ws_instr.row_dimensions[row].height = 22
            lc = ws_instr.cell(row=row, column=1, value=f"  {line}" if line else "")
            lc.font = body_font
            lc.alignment = left_center
            lc.border = thin_border
            ws_instr.cell(row=row, column=2).border = thin_border
            row += 1
    row += 1  # gap

# Freeze & print
ws_instr.freeze_panes = "A3"
ws_instr.print_area = f"A1:B{row}"

# ── SHEET 3: All Subscriptions (build FIRST so Dashboard can reference) ──
ws_subs = wb.create_sheet("All Subscriptions")
ws_subs.sheet_properties.tabColor = EMERALD

HEADERS_SUBS = [
    "Service Name", "Category", "Monthly Cost", "Annual Cost",
    "Billing Cycle", "Next Renewal Date", "Auto-Renew",
    "Payment Method", "Status", "Notes"
]
COL_WIDTHS_SUBS = [24, 20, 15, 15, 16, 18, 14, 18, 20, 30]

# Title row
ws_subs.merge_cells("A1:J1")
ws_subs.row_dimensions[1].height = 42
tc = ws_subs["A1"]
tc.value = "ALL SUBSCRIPTIONS"
tc.font = Font(name="Aptos", bold=True, color=WHITE, size=18)
tc.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
tc.alignment = Alignment(horizontal="center", vertical="center")
for col_i in range(1, 11):
    ws_subs.cell(row=1, column=col_i).fill = PatternFill("solid", fgColor=DEEP_PURPLE)
    ws_subs.cell(row=1, column=col_i).border = medium_border

# Subtitle
ws_subs.merge_cells("A2:J2")
ws_subs.row_dimensions[2].height = 26
sc = ws_subs["A2"]
sc.value = "Enter all your subscriptions below. Use the dropdowns for categories, billing cycles, and status."
sc.font = Font(name="Aptos", italic=True, color=MID_PURPLE, size=10)
sc.fill = PatternFill("solid", fgColor="EDE9FE")
sc.alignment = Alignment(horizontal="center", vertical="center")

# Headers row 3
for i, h in enumerate(HEADERS_SUBS, 1):
    ws_subs.column_dimensions[get_column_letter(i)].width = COL_WIDTHS_SUBS[i - 1]

style_header_row(ws_subs, 3, 10, height=30)
for i, h in enumerate(HEADERS_SUBS, 1):
    ws_subs.cell(row=3, column=i, value=h)
    ws_subs.cell(row=3, column=i).font = header_font
    ws_subs.cell(row=3, column=i).fill = header_fill
    ws_subs.cell(row=3, column=i).alignment = center
    ws_subs.cell(row=3, column=i).border = medium_border

# Data rows 4-103 (100 rows)
DATA_START = 4
DATA_END = 103

for r in range(DATA_START, DATA_END + 1):
    ws_subs.row_dimensions[r].height = 26
    for c in range(1, 11):
        cell = ws_subs.cell(row=r, column=c)
        is_curr = c in (3, 4)
        is_dt = c == 6
        is_ctr = c in (5, 7, 9)
        style_data_cell(cell, r, is_currency=is_curr, is_date=is_dt, is_center=is_ctr)

    # Annual Cost formula (col D = col C * 12 when billing is Monthly, else direct)
    formula = (
        f'=IF(C{r}="","",IF(E{r}="Annual",C{r}*12,'
        f'IF(E{r}="Quarterly",C{r}*4,'
        f'IF(E{r}="Weekly",C{r}*52,C{r}*12))))'
    )
    ws_subs.cell(row=r, column=4).value = formula
    ws_subs.cell(row=r, column=4).number_format = currency_fmt
    ws_subs.cell(row=r, column=4).protection = Protection(locked=True)

# Dropdowns
dv_cat = DataValidation(type="list", formula1='"' + ",".join(CATEGORIES) + '"', allow_blank=True)
dv_cat.error = "Please select a valid category"
dv_cat.errorTitle = "Invalid Category"
dv_cat.prompt = "Select a category"
dv_cat.promptTitle = "Category"
ws_subs.add_data_validation(dv_cat)
dv_cat.add(f"B{DATA_START}:B{DATA_END}")

dv_cycle = DataValidation(type="list", formula1='"Monthly,Annual,Quarterly,Weekly"', allow_blank=True)
ws_subs.add_data_validation(dv_cycle)
dv_cycle.add(f"E{DATA_START}:E{DATA_END}")

dv_renew = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_subs.add_data_validation(dv_renew)
dv_renew.add(f"G{DATA_START}:G{DATA_END}")

dv_pay = DataValidation(type="list",
    formula1='"Credit Card,Debit,PayPal,Apple Pay,Google Pay,Bank Transfer"',
    allow_blank=True)
ws_subs.add_data_validation(dv_pay)
dv_pay.add(f"H{DATA_START}:H{DATA_END}")

dv_status = DataValidation(type="list",
    formula1='"Active,Paused,Consider Canceling,Canceled"',
    allow_blank=True)
ws_subs.add_data_validation(dv_status)
dv_status.add(f"I{DATA_START}:I{DATA_END}")

# Conditional formatting on Status column (I)
green_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
yellow_fill = PatternFill("solid", fgColor=LIGHT_AMBER)
orange_fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
strikethrough_font = Font(name="Aptos", color=GRAY, size=11, strikethrough=True)

status_range = f"I{DATA_START}:I{DATA_END}"
ws_subs.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Active"'], fill=green_fill,
              font=Font(name="Aptos", color=EMERALD, bold=True, size=11)))
ws_subs.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Paused"'], fill=yellow_fill,
              font=Font(name="Aptos", color=AMBER, bold=True, size=11)))
ws_subs.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Consider Canceling"'], fill=orange_fill,
              font=Font(name="Aptos", color=ORANGE, bold=True, size=11)))
ws_subs.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Canceled"'], fill=gray_fill,
              font=strikethrough_font))

# Also apply row-level conditional formatting for Canceled rows (strikethrough entire row)
for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "J"]:
    row_range = f"{col_letter}{DATA_START}:{col_letter}{DATA_END}"
    ws_subs.conditional_formatting.add(row_range,
        FormulaRule(formula=[f'$I{DATA_START}="Canceled"'], fill=gray_fill, font=strikethrough_font))

ws_subs.freeze_panes = "A4"
ws_subs.auto_filter.ref = f"A3:J{DATA_END}"
ws_subs.page_margins = openpyxl.worksheet.page.PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)
ws_subs.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

# ── SHEET 2: Dashboard ────────────────────────────────────────
ws_dash = wb.create_sheet("Dashboard")
wb.move_sheet("Dashboard", offset=-1)  # move before All Subscriptions
ws_dash.sheet_properties.tabColor = MID_PURPLE

ws_dash.column_dimensions["A"].width = 4
ws_dash.column_dimensions["B"].width = 30
ws_dash.column_dimensions["C"].width = 18
ws_dash.column_dimensions["D"].width = 4
ws_dash.column_dimensions["E"].width = 30
ws_dash.column_dimensions["F"].width = 18
ws_dash.column_dimensions["G"].width = 18
ws_dash.column_dimensions["H"].width = 18

# Title
ws_dash.merge_cells("A1:H1")
ws_dash.row_dimensions[1].height = 50
t = ws_dash["A1"]
t.value = "SUBSCRIPTION DASHBOARD"
t.font = Font(name="Aptos", bold=True, color=WHITE, size=22)
t.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
t.alignment = Alignment(horizontal="center", vertical="center")
for ci in range(1, 9):
    ws_dash.cell(row=1, column=ci).fill = PatternFill("solid", fgColor=DEEP_PURPLE)
    ws_dash.cell(row=1, column=ci).border = medium_border

ws_dash.merge_cells("A2:H2")
ws_dash.row_dimensions[2].height = 28
s2 = ws_dash["A2"]
s2.value = "All figures update automatically from the All Subscriptions sheet."
s2.font = Font(name="Aptos", italic=True, color=MID_PURPLE, size=10)
s2.fill = PatternFill("solid", fgColor="EDE9FE")
s2.alignment = Alignment(horizontal="center", vertical="center")

# ── KEY METRICS ──
row = 4
write_section_header(ws_dash, row, 2, 3, "KEY METRICS")
write_section_header(ws_dash, row, 5, 6, "QUICK STATS")
row += 1

# Total Monthly Cost
vc = write_kpi_box(ws_dash, row, 2, "Total Monthly Cost",
    '=SUMPRODUCT(('
    f"'All Subscriptions'!I{DATA_START}:I{DATA_END}<>\"Canceled\")*"
    f"'All Subscriptions'!C{DATA_START}:C{DATA_END})")
vc.number_format = currency_fmt
# Active subscriptions count
vc2 = write_kpi_box(ws_dash, row, 5, "Active Subscriptions",
    f"=COUNTIF('All Subscriptions'!I{DATA_START}:I{DATA_END},\"Active\")")
vc2.number_format = "0"
row += 1

# Total Annual Cost
vc = write_kpi_box(ws_dash, row, 2, "Total Annual Cost",
    '=SUMPRODUCT(('
    f"'All Subscriptions'!I{DATA_START}:I{DATA_END}<>\"Canceled\")*"
    f"'All Subscriptions'!D{DATA_START}:D{DATA_END})")
vc.number_format = currency_fmt
# Paused count
vc2 = write_kpi_box(ws_dash, row, 5, "Paused Subscriptions",
    f"=COUNTIF('All Subscriptions'!I{DATA_START}:I{DATA_END},\"Paused\")")
vc2.number_format = "0"
row += 1

# Average Monthly Cost
vc = write_kpi_box(ws_dash, row, 2, "Average Monthly Cost",
    '=IFERROR(SUMPRODUCT(('
    f"'All Subscriptions'!I{DATA_START}:I{DATA_END}<>\"Canceled\")*"
    f"'All Subscriptions'!C{DATA_START}:C{DATA_END})"
    f"/COUNTIF('All Subscriptions'!I{DATA_START}:I{DATA_END},\"Active\"),0)")
vc.number_format = currency_fmt
# Consider canceling count
vc2 = write_kpi_box(ws_dash, row, 5, "Consider Canceling",
    f"=COUNTIF('All Subscriptions'!I{DATA_START}:I{DATA_END},\"Consider Canceling\")")
vc2.number_format = "0"
row += 1

# Most Expensive
vc = write_kpi_box(ws_dash, row, 2, "Most Expensive (Monthly)",
    f"=IFERROR(MAX('All Subscriptions'!C{DATA_START}:C{DATA_END}),0)")
vc.number_format = currency_fmt
# Canceled count
vc2 = write_kpi_box(ws_dash, row, 5, "Canceled Subscriptions",
    f"=COUNTIF('All Subscriptions'!I{DATA_START}:I{DATA_END},\"Canceled\")")
vc2.number_format = "0"
row += 2

# ── CATEGORY BREAKDOWN ──
cat_start_row = row
write_section_header(ws_dash, row, 2, 6, "CATEGORY BREAKDOWN")
row += 1

# Sub-headers
for ci, hdr in [(2, "Category"), (3, "Count"), (5, "Monthly Total"), (6, "Annual Total")]:
    cell = ws_dash.cell(row=row, column=ci, value=hdr)
    cell.font = Font(name="Aptos", bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
    cell.alignment = center
    cell.border = medium_border
# spacer col D
ws_dash.cell(row=row, column=4).fill = PatternFill("solid", fgColor=DEEP_PURPLE)
ws_dash.cell(row=row, column=4).border = medium_border
row += 1

for idx, cat in enumerate(CATEGORIES):
    bg, fg = CAT_COLORS[cat]
    r = row + idx
    ws_dash.row_dimensions[r].height = 26

    # Category name
    cell = ws_dash.cell(row=r, column=2, value=cat)
    cell.font = Font(name="Aptos", bold=True, color=fg, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = thin_border
    cell.alignment = left_center

    # Count
    cell = ws_dash.cell(row=r, column=3,
        value=f'=COUNTIFS(\'All Subscriptions\'!B{DATA_START}:B{DATA_END},"{cat}",'
              f"'All Subscriptions'!I{DATA_START}:I{DATA_END},\"<>Canceled\")")
    cell.font = Font(name="Aptos", bold=True, color=DARK_TEXT, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = thin_border
    cell.alignment = center
    cell.number_format = "0"

    # spacer col D
    ws_dash.cell(row=r, column=4).fill = PatternFill("solid", fgColor=WHITE)

    # Monthly Total
    cell = ws_dash.cell(row=r, column=5,
        value='=SUMPRODUCT(('
              f"'All Subscriptions'!B{DATA_START}:B{DATA_END}=\"{cat}\")*"
              f"('All Subscriptions'!I{DATA_START}:I{DATA_END}<>\"Canceled\")*"
              f"'All Subscriptions'!C{DATA_START}:C{DATA_END})")
    cell.font = Font(name="Aptos", color=DARK_TEXT, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = thin_border
    cell.alignment = right_center
    cell.number_format = currency_fmt

    # Annual Total
    cell = ws_dash.cell(row=r, column=6,
        value='=SUMPRODUCT(('
              f"'All Subscriptions'!B{DATA_START}:B{DATA_END}=\"{cat}\")*"
              f"('All Subscriptions'!I{DATA_START}:I{DATA_END}<>\"Canceled\")*"
              f"'All Subscriptions'!D{DATA_START}:D{DATA_END})")
    cell.font = Font(name="Aptos", color=DARK_TEXT, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = thin_border
    cell.alignment = right_center
    cell.number_format = currency_fmt

# Conditional formatting: monthly > $50 = amber highlight
cat_monthly_range = f"E{row}:E{row + len(CATEGORIES) - 1}"
ws_dash.conditional_formatting.add(cat_monthly_range,
    CellIsRule(operator="greaterThan", formula=["50"],
              fill=PatternFill("solid", fgColor="FDE68A"),
              font=Font(name="Aptos", bold=True, color=AMBER, size=11)))

row += len(CATEGORIES)

# Totals row
ws_dash.row_dimensions[row].height = 28
cell = ws_dash.cell(row=row, column=2, value="TOTAL")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=12)
cell.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
cell.border = medium_border
cell.alignment = left_center

cell = ws_dash.cell(row=row, column=3,
    value=f"=SUM(C{row - len(CATEGORIES)}:C{row - 1})")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=12)
cell.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
cell.border = medium_border
cell.alignment = center
cell.number_format = "0"

ws_dash.cell(row=row, column=4).fill = PatternFill("solid", fgColor=DEEP_PURPLE)
ws_dash.cell(row=row, column=4).border = medium_border

cell = ws_dash.cell(row=row, column=5,
    value=f"=SUM(E{row - len(CATEGORIES)}:E{row - 1})")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=12)
cell.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
cell.border = medium_border
cell.alignment = right_center
cell.number_format = currency_fmt

cell = ws_dash.cell(row=row, column=6,
    value=f"=SUM(F{row - len(CATEGORIES)}:F{row - 1})")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=12)
cell.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
cell.border = medium_border
cell.alignment = right_center
cell.number_format = currency_fmt

row += 2

# ── POTENTIAL SAVINGS ──
write_section_header(ws_dash, row, 2, 6, 'POTENTIAL SAVINGS -- Subscriptions Marked "Consider Canceling"')
row += 1

# Sub-headers
for ci, hdr in [(2, "Service Name"), (3, "Category"), (4, ""), (5, "Monthly Cost"), (6, "Annual Cost")]:
    cell = ws_dash.cell(row=row, column=ci, value=hdr)
    cell.font = Font(name="Aptos", bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=ORANGE)
    cell.alignment = center
    cell.border = medium_border
row += 1

# Use SMALL/IF array formulas to list up to 15 "Consider Canceling" items
savings_start = row
for i in range(1, 16):
    r = row + i - 1
    ws_dash.row_dimensions[r].height = 24

    alt_fill = PatternFill("solid", fgColor=LIGHT_ORANGE if i % 2 else WHITE)

    # Service Name
    cell = ws_dash.cell(row=r, column=2,
        value=f'=IFERROR(INDEX(\'All Subscriptions\'!A{DATA_START}:A{DATA_END},'
              f'SMALL(IF(\'All Subscriptions\'!I{DATA_START}:I{DATA_END}="Consider Canceling",'
              f'ROW(\'All Subscriptions\'!I{DATA_START}:I{DATA_END})-{DATA_START-1}),{i})),"")')
    cell.font = body_font
    cell.fill = alt_fill
    cell.border = thin_border
    cell.alignment = left_center

    # Category
    cell = ws_dash.cell(row=r, column=3,
        value=f'=IFERROR(INDEX(\'All Subscriptions\'!B{DATA_START}:B{DATA_END},'
              f'SMALL(IF(\'All Subscriptions\'!I{DATA_START}:I{DATA_END}="Consider Canceling",'
              f'ROW(\'All Subscriptions\'!I{DATA_START}:I{DATA_END})-{DATA_START-1}),{i})),"")')
    cell.font = body_font
    cell.fill = alt_fill
    cell.border = thin_border
    cell.alignment = center

    ws_dash.cell(row=r, column=4).fill = PatternFill("solid", fgColor=WHITE)

    # Monthly Cost
    cell = ws_dash.cell(row=r, column=5,
        value=f'=IFERROR(INDEX(\'All Subscriptions\'!C{DATA_START}:C{DATA_END},'
              f'SMALL(IF(\'All Subscriptions\'!I{DATA_START}:I{DATA_END}="Consider Canceling",'
              f'ROW(\'All Subscriptions\'!I{DATA_START}:I{DATA_END})-{DATA_START-1}),{i})),"")')
    cell.font = body_font
    cell.fill = alt_fill
    cell.border = thin_border
    cell.alignment = right_center
    cell.number_format = currency_fmt

    # Annual Cost
    cell = ws_dash.cell(row=r, column=6,
        value=f'=IFERROR(INDEX(\'All Subscriptions\'!D{DATA_START}:D{DATA_END},'
              f'SMALL(IF(\'All Subscriptions\'!I{DATA_START}:I{DATA_END}="Consider Canceling",'
              f'ROW(\'All Subscriptions\'!I{DATA_START}:I{DATA_END})-{DATA_START-1}),{i})),"")')
    cell.font = body_font
    cell.fill = alt_fill
    cell.border = thin_border
    cell.alignment = right_center
    cell.number_format = currency_fmt

row += 15

# Savings total
row += 1
ws_dash.row_dimensions[row].height = 30
cell = ws_dash.cell(row=row, column=2, value="TOTAL POTENTIAL SAVINGS")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=12)
cell.fill = PatternFill("solid", fgColor=RED)
cell.border = medium_border
cell.alignment = left_center

for ci in [3, 4]:
    ws_dash.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=RED)
    ws_dash.cell(row=row, column=ci).border = medium_border

cell = ws_dash.cell(row=row, column=5,
    value='=SUMPRODUCT(('
          f"'All Subscriptions'!I{DATA_START}:I{DATA_END}=\"Consider Canceling\")*"
          f"'All Subscriptions'!C{DATA_START}:C{DATA_END})")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=14)
cell.fill = PatternFill("solid", fgColor=RED)
cell.border = medium_border
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.number_format = currency_fmt

cell = ws_dash.cell(row=row, column=6,
    value='=SUMPRODUCT(('
          f"'All Subscriptions'!I{DATA_START}:I{DATA_END}=\"Consider Canceling\")*"
          f"'All Subscriptions'!D{DATA_START}:D{DATA_END})")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=14)
cell.fill = PatternFill("solid", fgColor=RED)
cell.border = medium_border
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.number_format = currency_fmt

ws_dash.freeze_panes = "A3"
ws_dash.page_margins = openpyxl.worksheet.page.PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)

# ── SHEET 4: Renewal Calendar ─────────────────────────────────
ws_cal = wb.create_sheet("Renewal Calendar")
ws_cal.sheet_properties.tabColor = AMBER

ws_cal.column_dimensions["A"].width = 3
ws_cal.column_dimensions["B"].width = 28
ws_cal.column_dimensions["C"].width = 18
ws_cal.column_dimensions["D"].width = 16
ws_cal.column_dimensions["E"].width = 18

# Title
ws_cal.merge_cells("A1:E1")
ws_cal.row_dimensions[1].height = 50
tc = ws_cal["A1"]
tc.value = "RENEWAL CALENDAR"
tc.font = Font(name="Aptos", bold=True, color=WHITE, size=22)
tc.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
tc.alignment = Alignment(horizontal="center", vertical="center")
for ci in range(1, 6):
    ws_cal.cell(row=1, column=ci).fill = PatternFill("solid", fgColor=DEEP_PURPLE)
    ws_cal.cell(row=1, column=ci).border = medium_border

ws_cal.merge_cells("A2:E2")
ws_cal.row_dimensions[2].height = 26
sc = ws_cal["A2"]
sc.value = "See when your subscriptions renew each month. Formulas pull from the All Subscriptions sheet."
sc.font = Font(name="Aptos", italic=True, color=MID_PURPLE, size=10)
sc.fill = PatternFill("solid", fgColor="EDE9FE")
sc.alignment = Alignment(horizontal="center", vertical="center")

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]

row = 4
for m_idx, month_name in enumerate(MONTHS):
    month_num = m_idx + 1

    # Month header
    ws_cal.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws_cal.row_dimensions[row].height = 32
    cell = ws_cal.cell(row=row, column=2, value=month_name.upper())
    cell.font = Font(name="Aptos", bold=True, color=WHITE, size=14)
    cell.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = medium_border
    for ci in range(3, 6):
        ws_cal.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=DEEP_PURPLE)
        ws_cal.cell(row=row, column=ci).border = medium_border
    row += 1

    # Column sub-headers
    for ci, hdr in [(2, "Service"), (3, "Category"), (4, "Renewal Date"), (5, "Amount")]:
        cell = ws_cal.cell(row=row, column=ci, value=hdr)
        cell.font = Font(name="Aptos", bold=True, color=MID_PURPLE, size=10)
        cell.fill = PatternFill("solid", fgColor="EDE9FE")
        cell.alignment = center
        cell.border = thin_border
    row += 1

    # Up to 12 entries per month
    entries_per_month = 12
    for i in range(1, entries_per_month + 1):
        r = row
        ws_cal.row_dimensions[r].height = 22

        alt_fill = PatternFill("solid", fgColor=LIGHT_PURPLE if i % 2 == 0 else WHITE)

        # Service Name
        cell = ws_cal.cell(row=r, column=2,
            value=f'=IFERROR(INDEX(\'All Subscriptions\'!A{DATA_START}:A{DATA_END},'
                  f'SMALL(IF(MONTH(\'All Subscriptions\'!F{DATA_START}:F{DATA_END})={month_num},'
                  f'ROW(\'All Subscriptions\'!F{DATA_START}:F{DATA_END})-{DATA_START-1}),{i})),"")')
        cell.font = body_font
        cell.fill = alt_fill
        cell.border = thin_border
        cell.alignment = left_center

        # Category
        cell = ws_cal.cell(row=r, column=3,
            value=f'=IFERROR(INDEX(\'All Subscriptions\'!B{DATA_START}:B{DATA_END},'
                  f'SMALL(IF(MONTH(\'All Subscriptions\'!F{DATA_START}:F{DATA_END})={month_num},'
                  f'ROW(\'All Subscriptions\'!F{DATA_START}:F{DATA_END})-{DATA_START-1}),{i})),"")')
        cell.font = body_font
        cell.fill = alt_fill
        cell.border = thin_border
        cell.alignment = center

        # Renewal Date
        cell = ws_cal.cell(row=r, column=4,
            value=f'=IFERROR(INDEX(\'All Subscriptions\'!F{DATA_START}:F{DATA_END},'
                  f'SMALL(IF(MONTH(\'All Subscriptions\'!F{DATA_START}:F{DATA_END})={month_num},'
                  f'ROW(\'All Subscriptions\'!F{DATA_START}:F{DATA_END})-{DATA_START-1}),{i})),"")')
        cell.font = body_font
        cell.fill = alt_fill
        cell.border = thin_border
        cell.alignment = center
        cell.number_format = date_fmt

        # Amount
        cell = ws_cal.cell(row=r, column=5,
            value=f'=IFERROR(INDEX(\'All Subscriptions\'!C{DATA_START}:C{DATA_END},'
                  f'SMALL(IF(MONTH(\'All Subscriptions\'!F{DATA_START}:F{DATA_END})={month_num},'
                  f'ROW(\'All Subscriptions\'!F{DATA_START}:F{DATA_END})-{DATA_START-1}),{i})),"")')
        cell.font = body_font
        cell.fill = alt_fill
        cell.border = thin_border
        cell.alignment = right_center
        cell.number_format = currency_fmt

        row += 1

    # Monthly total row
    ws_cal.row_dimensions[row].height = 28
    cell = ws_cal.cell(row=row, column=2, value=f"{month_name} Total")
    cell.font = Font(name="Aptos", bold=True, color=DEEP_PURPLE, size=11)
    cell.fill = PatternFill("solid", fgColor="EDE9FE")
    cell.border = medium_border
    cell.alignment = left_center

    for ci in [3, 4]:
        ws_cal.cell(row=row, column=ci).fill = PatternFill("solid", fgColor="EDE9FE")
        ws_cal.cell(row=row, column=ci).border = medium_border

    cell = ws_cal.cell(row=row, column=5,
        value=f'=SUMPRODUCT((MONTH(\'All Subscriptions\'!F{DATA_START}:F{DATA_END})={month_num})*'
              f"('All Subscriptions'!I{DATA_START}:I{DATA_END}<>\"Canceled\")*"
              f"'All Subscriptions'!C{DATA_START}:C{DATA_END})")
    cell.font = Font(name="Aptos", bold=True, color=DEEP_PURPLE, size=12)
    cell.fill = PatternFill("solid", fgColor="EDE9FE")
    cell.border = medium_border
    cell.alignment = right_center
    cell.number_format = currency_fmt

    row += 2  # gap between months

ws_cal.freeze_panes = "A3"
ws_cal.page_margins = openpyxl.worksheet.page.PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)

# ── SHEET 5: Annual Summary ───────────────────────────────────
ws_annual = wb.create_sheet("Annual Summary")
ws_annual.sheet_properties.tabColor = "059669"

ws_annual.column_dimensions["A"].width = 3
ws_annual.column_dimensions["B"].width = 22
for ci in range(3, 15):
    ws_annual.column_dimensions[get_column_letter(ci)].width = 14
ws_annual.column_dimensions["O"].width = 16  # Total column

# Title
ws_annual.merge_cells("A1:O1")
ws_annual.row_dimensions[1].height = 50
tc = ws_annual["A1"]
tc.value = "ANNUAL SUMMARY"
tc.font = Font(name="Aptos", bold=True, color=WHITE, size=22)
tc.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
tc.alignment = Alignment(horizontal="center", vertical="center")
for ci in range(1, 16):
    ws_annual.cell(row=1, column=ci).fill = PatternFill("solid", fgColor=DEEP_PURPLE)
    ws_annual.cell(row=1, column=ci).border = medium_border

ws_annual.merge_cells("A2:O2")
ws_annual.row_dimensions[2].height = 26
sc = ws_annual["A2"]
sc.value = "Track your subscription spending month by month. Enter actual monthly totals below."
sc.font = Font(name="Aptos", italic=True, color=MID_PURPLE, size=10)
sc.fill = PatternFill("solid", fgColor="EDE9FE")
sc.alignment = Alignment(horizontal="center", vertical="center")

# ── MONTH-BY-MONTH SPENDING ──
row = 4
write_section_header(ws_annual, row, 2, 15, "MONTH-BY-MONTH SPENDING")
row += 1

# Headers
month_abbrevs = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
all_headers = [""] + month_abbrevs + ["TOTAL"]
for ci, hdr in enumerate(all_headers, 1):
    if ci == 1:
        continue
    cell = ws_annual.cell(row=row, column=ci, value=hdr)
    cell.font = Font(name="Aptos", bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
    cell.alignment = center
    cell.border = medium_border
row += 1

# Current Year row
yr_row_current = row
ws_annual.row_dimensions[row].height = 28
cell = ws_annual.cell(row=row, column=2, value="Current Year")
cell.font = Font(name="Aptos", bold=True, color=DEEP_PURPLE, size=11)
cell.fill = PatternFill("solid", fgColor="EDE9FE")
cell.border = medium_border
cell.alignment = left_center

for ci in range(3, 15):
    cell = ws_annual.cell(row=row, column=ci)
    cell.number_format = currency_fmt
    cell.font = body_font
    cell.fill = PatternFill("solid", fgColor=WHITE)
    cell.border = thin_border
    cell.alignment = right_center

# Total
cell = ws_annual.cell(row=row, column=15,
    value=f"=SUM(C{row}:N{row})")
cell.font = Font(name="Aptos", bold=True, color=DEEP_PURPLE, size=12)
cell.fill = PatternFill("solid", fgColor="EDE9FE")
cell.border = medium_border
cell.alignment = right_center
cell.number_format = currency_fmt
row += 1

# Previous Year row
yr_row_prev = row
ws_annual.row_dimensions[row].height = 28
cell = ws_annual.cell(row=row, column=2, value="Previous Year")
cell.font = Font(name="Aptos", bold=True, color=GRAY, size=11)
cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
cell.border = medium_border
cell.alignment = left_center

for ci in range(3, 15):
    cell = ws_annual.cell(row=row, column=ci)
    cell.number_format = currency_fmt
    cell.font = Font(name="Aptos", color=GRAY, size=11)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.border = thin_border
    cell.alignment = right_center

cell = ws_annual.cell(row=row, column=15,
    value=f"=SUM(C{row}:N{row})")
cell.font = Font(name="Aptos", bold=True, color=GRAY, size=12)
cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
cell.border = medium_border
cell.alignment = right_center
cell.number_format = currency_fmt
row += 1

# Difference row
diff_row = row
ws_annual.row_dimensions[row].height = 28
cell = ws_annual.cell(row=row, column=2, value="Difference (+/-)")
cell.font = Font(name="Aptos", bold=True, color=DARK_TEXT, size=11)
cell.fill = PatternFill("solid", fgColor="FEF3C7")
cell.border = medium_border
cell.alignment = left_center

for ci in range(3, 15):
    cell = ws_annual.cell(row=row, column=ci,
        value=f"={get_column_letter(ci)}{yr_row_current}-{get_column_letter(ci)}{yr_row_prev}")
    cell.number_format = currency_fmt
    cell.font = Font(name="Aptos", color=DARK_TEXT, size=11)
    cell.fill = PatternFill("solid", fgColor="FEF3C7")
    cell.border = thin_border
    cell.alignment = right_center

cell = ws_annual.cell(row=row, column=15,
    value=f"=O{yr_row_current}-O{yr_row_prev}")
cell.font = Font(name="Aptos", bold=True, color=DARK_TEXT, size=12)
cell.fill = PatternFill("solid", fgColor="FEF3C7")
cell.border = medium_border
cell.alignment = right_center
cell.number_format = currency_fmt

# Conditional formatting: positive difference = red, negative = green (saving money)
diff_range = f"C{diff_row}:O{diff_row}"
ws_annual.conditional_formatting.add(diff_range,
    CellIsRule(operator="greaterThan", formula=["0"],
              font=Font(name="Aptos", color=RED, bold=True, size=11)))
ws_annual.conditional_formatting.add(diff_range,
    CellIsRule(operator="lessThan", formula=["0"],
              font=Font(name="Aptos", color=EMERALD, bold=True, size=11)))

row += 2

# ── CATEGORY-BY-CATEGORY BREAKDOWN ──
write_section_header(ws_annual, row, 2, 15, "CATEGORY-BY-CATEGORY ANNUAL BREAKDOWN")
row += 1

# Headers
for ci, hdr in enumerate(all_headers, 1):
    if ci == 1:
        continue
    cell = ws_annual.cell(row=row, column=ci, value=hdr)
    cell.font = Font(name="Aptos", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=MID_PURPLE)
    cell.alignment = center
    cell.border = medium_border
row += 1

cat_data_start = row
for idx, cat in enumerate(CATEGORIES):
    r = row + idx
    bg, fg = CAT_COLORS[cat]
    ws_annual.row_dimensions[r].height = 26

    cell = ws_annual.cell(row=r, column=2, value=cat)
    cell.font = Font(name="Aptos", bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = thin_border
    cell.alignment = left_center

    for ci in range(3, 15):
        cell = ws_annual.cell(row=r, column=ci)
        cell.number_format = currency_fmt
        cell.font = Font(name="Aptos", color=DARK_TEXT, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = thin_border
        cell.alignment = right_center

    # Total for category
    cell = ws_annual.cell(row=r, column=15,
        value=f"=SUM(C{r}:N{r})")
    cell.font = Font(name="Aptos", bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = medium_border
    cell.alignment = right_center
    cell.number_format = currency_fmt

row += len(CATEGORIES)

# Category total row
ws_annual.row_dimensions[row].height = 28
cell = ws_annual.cell(row=row, column=2, value="TOTAL")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=11)
cell.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
cell.border = medium_border
cell.alignment = left_center

for ci in range(3, 16):
    cell = ws_annual.cell(row=row, column=ci,
        value=f"=SUM({get_column_letter(ci)}{cat_data_start}:{get_column_letter(ci)}{row - 1})")
    cell.font = Font(name="Aptos", bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
    cell.border = medium_border
    cell.alignment = right_center
    cell.number_format = currency_fmt

row += 2

# ── SAVINGS FROM CANCELLATIONS ──
write_section_header(ws_annual, row, 2, 6, "TOTAL SAVED FROM CANCELLATIONS")
row += 1

vc = write_kpi_box(ws_annual, row, 2, "Monthly Savings (from log)",
    "=SUM('Cancellation Log'!C5:C104)")
vc.number_format = currency_fmt
row += 1

vc = write_kpi_box(ws_annual, row, 2, "Annual Savings (from log)",
    "=SUM('Cancellation Log'!D5:D104)")
vc.number_format = currency_fmt

ws_annual.freeze_panes = "B3"
ws_annual.page_margins = openpyxl.worksheet.page.PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)

# ── SHEET 6: Cancellation Log ─────────────────────────────────
ws_cancel = wb.create_sheet("Cancellation Log")
ws_cancel.sheet_properties.tabColor = RED

HEADERS_CANCEL = [
    "Service Name", "Date Canceled", "Monthly Savings",
    "Annual Savings", "Reason", "Would Re-subscribe?"
]
COL_W_CANCEL = [24, 16, 16, 16, 36, 20]

# Title row
ws_cancel.merge_cells("A1:F1")
ws_cancel.row_dimensions[1].height = 50
tc = ws_cancel["A1"]
tc.value = "CANCELLATION LOG"
tc.font = Font(name="Aptos", bold=True, color=WHITE, size=22)
tc.fill = PatternFill("solid", fgColor=DEEP_PURPLE)
tc.alignment = Alignment(horizontal="center", vertical="center")
for ci in range(1, 7):
    ws_cancel.cell(row=1, column=ci).fill = PatternFill("solid", fgColor=DEEP_PURPLE)
    ws_cancel.cell(row=1, column=ci).border = medium_border

# Motivational banner
ws_cancel.merge_cells("A2:F2")
ws_cancel.row_dimensions[2].height = 36
mc = ws_cancel["A2"]
mc.value = "Every subscription you cancel puts money back in your pocket. Track your wins below!"
mc.font = Font(name="Aptos", italic=True, color=MID_PURPLE, size=11)
mc.fill = PatternFill("solid", fgColor="EDE9FE")
mc.alignment = Alignment(horizontal="center", vertical="center")

# Savings summary row
ws_cancel.merge_cells("A3:B3")
ws_cancel.row_dimensions[3].height = 40
sc = ws_cancel["A3"]
sc.value = "YOU'VE SAVED THIS YEAR:"
sc.font = Font(name="Aptos", bold=True, color=WHITE, size=14)
sc.fill = PatternFill("solid", fgColor=EMERALD)
sc.alignment = Alignment(horizontal="right", vertical="center")
sc.border = medium_border
ws_cancel["B3"].fill = PatternFill("solid", fgColor=EMERALD)
ws_cancel["B3"].border = medium_border

# Monthly savings total
cell = ws_cancel.cell(row=3, column=3, value="=SUM(C5:C104)")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=16)
cell.fill = PatternFill("solid", fgColor=EMERALD)
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.border = medium_border
cell.number_format = '"$"#,##0.00"/mo"'

cell = ws_cancel.cell(row=3, column=4, value="=SUM(D5:D104)")
cell.font = Font(name="Aptos", bold=True, color=WHITE, size=16)
cell.fill = PatternFill("solid", fgColor=EMERALD)
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.border = medium_border
cell.number_format = '"$"#,##0.00"/yr"'

for ci in [5, 6]:
    ws_cancel.cell(row=3, column=ci).fill = PatternFill("solid", fgColor=EMERALD)
    ws_cancel.cell(row=3, column=ci).border = medium_border

# Headers
for i, h in enumerate(HEADERS_CANCEL, 1):
    ws_cancel.column_dimensions[get_column_letter(i)].width = COL_W_CANCEL[i - 1]

style_header_row(ws_cancel, 4, 6, height=30)
for i, h in enumerate(HEADERS_CANCEL, 1):
    c = ws_cancel.cell(row=4, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = medium_border

# Data rows 5-104 (100 rows)
CANCEL_START = 5
CANCEL_END = 104

for r in range(CANCEL_START, CANCEL_END + 1):
    ws_cancel.row_dimensions[r].height = 26
    for c_idx in range(1, 7):
        cell = ws_cancel.cell(row=r, column=c_idx)
        is_curr = c_idx in (3, 4)
        is_dt = c_idx == 2
        is_ctr = c_idx == 6
        style_data_cell(cell, r, is_currency=is_curr, is_date=is_dt, is_center=is_ctr)

    # Annual savings formula = monthly * 12
    ws_cancel.cell(row=r, column=4).value = f'=IF(C{r}="","",C{r}*12)'
    ws_cancel.cell(row=r, column=4).number_format = currency_fmt

# Dropdown for Would Re-subscribe
dv_resub = DataValidation(type="list", formula1='"Yes,No,Maybe"', allow_blank=True)
ws_cancel.add_data_validation(dv_resub)
dv_resub.add(f"F{CANCEL_START}:F{CANCEL_END}")

ws_cancel.freeze_panes = "A5"
ws_cancel.auto_filter.ref = f"A4:F{CANCEL_END}"
ws_cancel.page_margins = openpyxl.worksheet.page.PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)

# ═══════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════
OUTPUT = "/Users/timdunn/mobile_app_ideas/etsy-templates/subscription-tracker.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")

import os
size = os.path.getsize(OUTPUT)
print(f"File size: {size:,} bytes ({size/1024:.1f} KB)")
