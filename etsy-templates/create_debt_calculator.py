#!/usr/bin/env python3
"""
Debt Payoff Calculator - Premium Etsy Template
Creates a professional .xlsx with Instructions, Dashboard, Debt Input,
Snowball Plan, Avalanche Plan, and Comparison sheets.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ── Colour palette ──────────────────────────────────────────────────
CHARCOAL   = "2D3748"
TEAL       = "0D9488"
TEAL_LIGHT = "CCFBF1"
RED        = "DC2626"
RED_LIGHT  = "FEE2E2"
GOLD       = "F59E0B"
GOLD_LIGHT = "FEF3C7"
WHITE      = "FFFFFF"
ALT_ROW    = "F7FAFC"
LIGHT_GRAY = "E2E8F0"
MED_GRAY   = "94A3B8"
DARK_TEXT   = "1A202C"
EDITABLE_BG = "FFFDE7"

# ── Reusable style helpers ──────────────────────────────────────────
header_font    = Font(name="Calibri", bold=True, color=WHITE, size=12)
header_fill    = PatternFill("solid", fgColor=CHARCOAL)
sub_header_font = Font(name="Calibri", bold=True, color=DARK_TEXT, size=11)
sub_header_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
teal_fill      = PatternFill("solid", fgColor=TEAL)
teal_font      = Font(name="Calibri", bold=True, color=WHITE, size=11)
teal_light_fill = PatternFill("solid", fgColor=TEAL_LIGHT)
red_fill       = PatternFill("solid", fgColor=RED)
red_font       = Font(name="Calibri", bold=True, color=WHITE, size=11)
red_light_fill = PatternFill("solid", fgColor=RED_LIGHT)
gold_fill      = PatternFill("solid", fgColor=GOLD)
gold_font      = Font(name="Calibri", bold=True, color=DARK_TEXT, size=11)
gold_light_fill = PatternFill("solid", fgColor=GOLD_LIGHT)
body_font      = Font(name="Calibri", size=11, color=DARK_TEXT)
body_font_bold = Font(name="Calibri", size=11, color=DARK_TEXT, bold=True)
small_font     = Font(name="Calibri", size=10, color=MED_GRAY)
alt_fill       = PatternFill("solid", fgColor=ALT_ROW)
editable_fill  = PatternFill("solid", fgColor=EDITABLE_BG)
no_fill        = PatternFill(fill_type=None)

thin_border = Border(
    left=Side(style="thin", color=LIGHT_GRAY),
    right=Side(style="thin", color=LIGHT_GRAY),
    top=Side(style="thin", color=LIGHT_GRAY),
    bottom=Side(style="thin", color=LIGHT_GRAY),
)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

CURRENCY_FMT = '$#,##0.00'
PCT_FMT = '0.00%'
INT_FMT = '#,##0'

MAX_DEBTS = 20
MAX_MONTHS = 120

wb = openpyxl.Workbook()


def style_header_row(ws, row, max_col, font=header_font, fill=header_fill):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border


def apply_alt_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill == no_fill or cell.fill.fgColor is None or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = alt_fill


def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def title_block(ws, row, col, text, merge_end_col, font_size=20, fill=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", bold=True, size=font_size, color=CHARCOAL)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if fill:
        cell.fill = fill
    if merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)


def subtitle_block(ws, row, col, text, merge_end_col):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", size=12, color=MED_GRAY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)


# ====================================================================
# SHEET 1 - INSTRUCTIONS
# ====================================================================
ws_instr = wb.active
ws_instr.title = "Instructions"
ws_instr.sheet_properties.tabColor = CHARCOAL

set_col_widths(ws_instr, {"A": 4, "B": 80, "C": 4})

r = 2
title_block(ws_instr, r, 2, "DEBT PAYOFF CALCULATOR", 2, 24)
r += 1
subtitle_block(ws_instr, r, 2, "Your step-by-step guide to becoming debt-free", 2)

r += 2
cell = ws_instr.cell(row=r, column=2, value="Welcome!")
cell.font = Font(name="Calibri", bold=True, size=14, color=TEAL)
r += 1
welcome_text = (
    "Congratulations on taking the first step toward financial freedom! "
    "This calculator will help you create a clear, actionable plan to pay off all of your debts. "
    "Simply enter your debts on the \"Debt Input\" sheet, set your extra monthly payment, "
    "and the calculator does the rest -- generating personalized Snowball and Avalanche payoff plans "
    "so you can compare strategies and choose the best path for you."
)
cell = ws_instr.cell(row=r, column=2, value=welcome_text)
cell.font = body_font
cell.alignment = left_wrap
ws_instr.row_dimensions[r].height = 70

r += 2
cell = ws_instr.cell(row=r, column=2, value="How to Use This Calculator")
cell.font = Font(name="Calibri", bold=True, size=14, color=TEAL)

steps = [
    ("Step 1:", "Go to the \"Debt Input\" sheet and enter each debt -- name, balance, APR, and minimum payment."),
    ("Step 2:", "On the \"Dashboard\" sheet, enter your Extra Monthly Payment (the highlighted yellow cell). Even $50/month makes a huge difference!"),
    ("Step 3:", "Review the \"Snowball Plan\" sheet to see your payoff schedule when tackling the smallest balance first."),
    ("Step 4:", "Review the \"Avalanche Plan\" sheet to see your payoff schedule when tackling the highest interest rate first."),
    ("Step 5:", "Check the \"Comparison\" sheet to see which method saves you the most money and time."),
    ("Step 6:", "Print your chosen plan and stick it on the fridge!"),
]
for step_title, step_desc in steps:
    r += 1
    cell = ws_instr.cell(row=r, column=2, value=f"{step_title}  {step_desc}")
    cell.font = body_font
    cell.alignment = left_wrap
    ws_instr.row_dimensions[r].height = 30

r += 2
cell = ws_instr.cell(row=r, column=2, value="Snowball Method vs. Avalanche Method")
cell.font = Font(name="Calibri", bold=True, size=14, color=TEAL)

r += 1
cell = ws_instr.cell(row=r, column=2,
    value="SNOWBALL METHOD  --  Pay off the smallest balance first, then roll that payment into the next smallest.")
cell.font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
cell.alignment = left_wrap
ws_instr.row_dimensions[r].height = 30

r += 1
snowball_details = (
    "Pros: Quick wins build momentum and motivation. You see debts disappear fast.\n"
    "Cons: You may pay slightly more in total interest over time.\n"
    "Best for: People who need psychological motivation to stay on track."
)
cell = ws_instr.cell(row=r, column=2, value=snowball_details)
cell.font = body_font
cell.alignment = left_wrap
ws_instr.row_dimensions[r].height = 60

r += 2
cell = ws_instr.cell(row=r, column=2,
    value="AVALANCHE METHOD  --  Pay off the highest interest rate first, saving the most money mathematically.")
cell.font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
cell.alignment = left_wrap
ws_instr.row_dimensions[r].height = 30

r += 1
avalanche_details = (
    "Pros: Minimizes total interest paid. The mathematically optimal strategy.\n"
    "Cons: It may take longer to fully pay off your first debt, which can feel slow.\n"
    "Best for: People who are disciplined and motivated by saving the most money."
)
cell = ws_instr.cell(row=r, column=2, value=avalanche_details)
cell.font = body_font
cell.alignment = left_wrap
ws_instr.row_dimensions[r].height = 60

r += 2
cell = ws_instr.cell(row=r, column=2, value="Tips for Getting Out of Debt")
cell.font = Font(name="Calibri", bold=True, size=14, color=TEAL)

tips = [
    "Automate your payments so you never miss a due date.",
    "Use the debt snowball/avalanche to stay focused -- do NOT spread extra payments across all debts.",
    "Build a small emergency fund ($500-$1,000) first so unexpected expenses don't derail your plan.",
    "Look for ways to increase income: side hustles, selling unused items, negotiating a raise.",
    "Call your creditors and ask for lower interest rates -- it works more often than you think!",
    "Track your progress monthly. Watching balances drop is incredibly motivating.",
    "Celebrate milestones! When you pay off a debt, reward yourself (within reason).",
    "Avoid taking on new debt while paying off existing debt.",
    "Consider balance transfer offers (0% APR) to reduce interest, but read the fine print.",
    "Stay patient. Becoming debt-free is a marathon, not a sprint.",
]
for i, tip in enumerate(tips, 1):
    r += 1
    cell = ws_instr.cell(row=r, column=2, value=f"  {i}.  {tip}")
    cell.font = body_font
    cell.alignment = left_wrap
    ws_instr.row_dimensions[r].height = 25

r += 2
cell = ws_instr.cell(row=r, column=2,
    value="NOTE: Yellow-highlighted cells are for your input. All other cells contain formulas -- please do not edit them.")
cell.font = Font(name="Calibri", bold=True, size=11, color=RED)
cell.fill = PatternFill("solid", fgColor=GOLD_LIGHT)
cell.alignment = left_wrap
ws_instr.row_dimensions[r].height = 30

ws_instr.sheet_view.showGridLines = False


# ====================================================================
# SHEET 3 - DEBT INPUT (create before Dashboard so Dashboard can ref)
# ====================================================================
ws_input = wb.create_sheet("Debt Input")
ws_input.sheet_properties.tabColor = TEAL

set_col_widths(ws_input, {
    "A": 4, "B": 28, "C": 18, "D": 22, "E": 18, "F": 20
})

title_block(ws_input, 1, 2, "DEBT PAYOFF CALCULATOR", 6, 20)
subtitle_block(ws_input, 2, 2, "Enter your debts below. Yellow cells are for your input.", 6)

HEADER_ROW_INPUT = 4
headers_input = ["Debt Name", "Current Balance", "Interest Rate (APR %)", "Minimum Payment", "Type"]
for ci, h in enumerate(headers_input, 2):
    cell = ws_input.cell(row=HEADER_ROW_INPUT, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

DATA_START = HEADER_ROW_INPUT + 1
DATA_END = DATA_START + MAX_DEBTS - 1  # rows 5-24

for i in range(MAX_DEBTS):
    r = DATA_START + i
    ws_input.row_dimensions[r].height = 22
    for c in range(2, 7):
        cell = ws_input.cell(row=r, column=c)
        cell.fill = editable_fill
        cell.border = thin_border
        cell.font = body_font
        cell.alignment = center
    ws_input.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws_input.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws_input.cell(row=r, column=4).number_format = '0.00%'

sample = [
    ("Credit Card A", 4500, 0.2199, 90, "Credit Card"),
    ("Student Loan", 12000, 0.055, 150, "Student Loan"),
    ("Car Loan", 8500, 0.069, 250, "Car Loan"),
    ("Personal Loan", 3000, 0.1099, 75, "Personal Loan"),
    ("Medical Bill", 1500, 0.0, 50, "Medical"),
]
for idx, (name, bal, apr, minp, typ) in enumerate(sample):
    r = DATA_START + idx
    ws_input.cell(row=r, column=2, value=name)
    ws_input.cell(row=r, column=3, value=bal)
    ws_input.cell(row=r, column=4, value=apr)
    ws_input.cell(row=r, column=5, value=minp)
    ws_input.cell(row=r, column=6, value=typ)

TOTALS_ROW_INPUT = DATA_END + 1  # row 25
ws_input.cell(row=TOTALS_ROW_INPUT, column=2, value="TOTALS").font = Font(name="Calibri", bold=True, size=12, color=WHITE)
ws_input.cell(row=TOTALS_ROW_INPUT, column=2).fill = PatternFill("solid", fgColor=CHARCOAL)
ws_input.cell(row=TOTALS_ROW_INPUT, column=2).alignment = center
for c in [3, 4, 5, 6]:
    cell = ws_input.cell(row=TOTALS_ROW_INPUT, column=c)
    cell.fill = PatternFill("solid", fgColor=CHARCOAL)
    cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    cell.alignment = center
    cell.border = thin_border

ws_input.cell(row=TOTALS_ROW_INPUT, column=3).value = f'=SUMIF(C{DATA_START}:C{DATA_END},">0")'
ws_input.cell(row=TOTALS_ROW_INPUT, column=3).number_format = CURRENCY_FMT
ws_input.cell(row=TOTALS_ROW_INPUT, column=4).value = (
    f'=IF(SUMIF(C{DATA_START}:C{DATA_END},">0")=0,0,'
    f'SUMPRODUCT(C{DATA_START}:C{DATA_END},D{DATA_START}:D{DATA_END})'
    f'/SUMIF(C{DATA_START}:C{DATA_END},">0"))'
)
ws_input.cell(row=TOTALS_ROW_INPUT, column=4).number_format = PCT_FMT
ws_input.cell(row=TOTALS_ROW_INPUT, column=5).value = f'=SUMIF(E{DATA_START}:E{DATA_END},">0")'
ws_input.cell(row=TOTALS_ROW_INPUT, column=5).number_format = CURRENCY_FMT

# Data validations
dv_apr = DataValidation(type="decimal", operator="between",
                        formula1="0", formula2="1",
                        errorTitle="Invalid Rate",
                        error="Enter a rate between 0% and 100%.",
                        promptTitle="Interest Rate",
                        prompt="Enter the APR as a decimal (e.g., 0.22 for 22%).")
dv_apr.showErrorMessage = True
dv_apr.showInputMessage = True
ws_input.add_data_validation(dv_apr)
dv_apr.add(f"D{DATA_START}:D{DATA_END}")

dv_pos = DataValidation(type="decimal", operator="greaterThanOrEqual",
                        formula1="0",
                        errorTitle="Invalid Amount",
                        error="Amount must be a positive number.")
dv_pos.showErrorMessage = True
ws_input.add_data_validation(dv_pos)
dv_pos.add(f"C{DATA_START}:C{DATA_END}")
dv_pos.add(f"E{DATA_START}:E{DATA_END}")

dv_type = DataValidation(type="list",
                         formula1='"Credit Card,Student Loan,Car Loan,Personal Loan,Medical,Mortgage,Other"')
dv_type.showDropDown = False
ws_input.add_data_validation(dv_type)
dv_type.add(f"F{DATA_START}:F{DATA_END}")

# Conditional formatting: highlight highest interest rate
ws_input.conditional_formatting.add(
    f"D{DATA_START}:D{DATA_END}",
    FormulaRule(
        formula=[f"AND(D{DATA_START}=MAX(D${DATA_START}:D${DATA_END}),D{DATA_START}>0)"],
        fill=PatternFill("solid", fgColor=RED_LIGHT),
        font=Font(bold=True, color=RED),
    )
)

apply_alt_rows(ws_input, DATA_START, DATA_END, 6)

ws_input.freeze_panes = "B5"
ws_input.sheet_view.showGridLines = False


# ====================================================================
# SHEET 2 - DASHBOARD
# ====================================================================
ws_dash = wb.create_sheet("Dashboard")
wb.move_sheet("Dashboard", offset=-1)
ws_dash.sheet_properties.tabColor = TEAL

set_col_widths(ws_dash, {
    "A": 4, "B": 32, "C": 22, "D": 6, "E": 32, "F": 22, "G": 4
})

title_block(ws_dash, 1, 2, "DEBT PAYOFF DASHBOARD", 6, 22)
subtitle_block(ws_dash, 2, 2, "Your complete debt snapshot at a glance", 6)

# Overview header
r = 4
for c in range(2, 7):
    cell = ws_dash.cell(row=r, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border
ws_dash.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws_dash.cell(row=r, column=2, value="OVERVIEW")

DI = "'Debt Input'"

def dash_row(ws, row, label, formula, fmt=CURRENCY_FMT, val_fill=None, val_font=None):
    cell_l = ws.cell(row=row, column=2, value=label)
    cell_l.font = body_font_bold
    cell_l.alignment = Alignment(horizontal="left", vertical="center")
    cell_l.border = thin_border
    cell_v = ws.cell(row=row, column=3, value=formula)
    cell_v.number_format = fmt
    cell_v.font = val_font or body_font
    cell_v.alignment = center
    cell_v.border = thin_border
    if val_fill:
        cell_v.fill = val_fill
    return cell_v

r = 5
dash_row(ws_dash, r, "Total Debt", f"={DI}!C{TOTALS_ROW_INPUT}", CURRENCY_FMT,
         val_fill=red_light_fill, val_font=Font(name="Calibri", bold=True, size=14, color=RED))

r = 6
dash_row(ws_dash, r, "Total Minimum Payments", f"={DI}!E{TOTALS_ROW_INPUT}", CURRENCY_FMT)

r = 7
dash_row(ws_dash, r, "Number of Debts",
         f'=COUNTA({DI}!B{DATA_START}:B{DATA_END})', INT_FMT)

r = 8
dash_row(ws_dash, r, "Weighted Avg Interest Rate", f"={DI}!D{TOTALS_ROW_INPUT}", PCT_FMT)

# Extra payment input
r = 10
for c in range(2, 7):
    cell = ws_dash.cell(row=r, column=c)
    cell.fill = gold_fill
    cell.font = gold_font
    cell.alignment = center
    cell.border = thin_border
ws_dash.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws_dash.cell(row=r, column=2, value="YOUR EXTRA MONTHLY PAYMENT")

r = 11
cell_l = ws_dash.cell(row=r, column=2, value="Extra Payment Per Month -->")
cell_l.font = Font(name="Calibri", bold=True, size=13, color=CHARCOAL)
cell_l.alignment = Alignment(horizontal="right", vertical="center")
cell_l.border = thin_border

EXTRA_PMT_CELL = "C11"
cell_v = ws_dash.cell(row=r, column=3, value=200)
cell_v.number_format = CURRENCY_FMT
cell_v.font = Font(name="Calibri", bold=True, size=16, color=TEAL)
cell_v.fill = editable_fill
cell_v.alignment = center
cell_v.border = Border(
    left=Side(style="thick", color=GOLD),
    right=Side(style="thick", color=GOLD),
    top=Side(style="thick", color=GOLD),
    bottom=Side(style="thick", color=GOLD),
)
cell_note = ws_dash.cell(row=r, column=4, value="<-- Enter your extra monthly payment here")
cell_note.font = Font(name="Calibri", italic=True, size=11, color=MED_GRAY)
cell_note.alignment = Alignment(horizontal="left", vertical="center")
ws_dash.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

dv_extra = DataValidation(type="decimal", operator="greaterThanOrEqual",
                          formula1="0", errorTitle="Invalid", error="Enter a positive number.")
dv_extra.showErrorMessage = True
ws_dash.add_data_validation(dv_extra)
dv_extra.add(EXTRA_PMT_CELL)

# Comparison section
r = 13
for c in range(2, 7):
    cell = ws_dash.cell(row=r, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border
ws_dash.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws_dash.cell(row=r, column=2, value="PAYOFF COMPARISON")

r = 14
for ci, txt in [(2, "Metric"), (3, "Snowball"), (5, "Avalanche")]:
    cell = ws_dash.cell(row=r, column=ci, value=txt)
    cell.font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.alignment = center
    cell.border = thin_border
# Merge 3-4 and 5-6 for sub headers
for c in [4, 6]:
    cell = ws_dash.cell(row=r, column=c)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.border = thin_border
ws_dash.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
ws_dash.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)

COMP = "'Comparison'"
comp_metrics = [
    ("Total Months to Payoff", f"={COMP}!C5", f"={COMP}!D5", INT_FMT),
    ("Total Interest Paid", f"={COMP}!C6", f"={COMP}!D6", CURRENCY_FMT),
    ("Total Amount Paid", f"={COMP}!C7", f"={COMP}!D7", CURRENCY_FMT),
]
for i, (label, snow_f, aval_f, fmt) in enumerate(comp_metrics):
    row = 15 + i
    ws_dash.cell(row=row, column=2, value=label).font = body_font_bold
    ws_dash.cell(row=row, column=2).border = thin_border
    ws_dash.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")

    for c in [3, 4, 5, 6]:
        ws_dash.cell(row=row, column=c).border = thin_border
    
    c_s = ws_dash.cell(row=row, column=3, value=snow_f)
    c_s.number_format = fmt
    c_s.font = body_font
    c_s.alignment = center
    ws_dash.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

    c_a = ws_dash.cell(row=row, column=5, value=aval_f)
    c_a.number_format = fmt
    c_a.font = body_font
    c_a.alignment = center
    ws_dash.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

# Savings row
r = 18
for c in range(2, 7):
    cell = ws_dash.cell(row=r, column=c)
    cell.fill = teal_light_fill
    cell.border = thin_border
ws_dash.cell(row=r, column=2, value="Interest Saved (Best Method)").font = Font(name="Calibri", bold=True, size=12, color=TEAL)
ws_dash.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_dash.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
c_saved = ws_dash.cell(row=r, column=3, value=f"={COMP}!C9")
c_saved.number_format = CURRENCY_FMT
c_saved.font = Font(name="Calibri", bold=True, size=14, color=TEAL)
c_saved.alignment = center

r = 19
for c in range(2, 7):
    cell = ws_dash.cell(row=r, column=c)
    cell.fill = gold_light_fill
    cell.border = thin_border
ws_dash.cell(row=r, column=2, value="Recommended Method").font = Font(name="Calibri", bold=True, size=12, color=CHARCOAL)
ws_dash.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_dash.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
ws_dash.cell(row=r, column=3, value=f"={COMP}!C10").font = Font(name="Calibri", bold=True, size=14, color=CHARCOAL)
ws_dash.cell(row=r, column=3).alignment = center

# Motivational section
r = 21
for c in range(2, 7):
    cell = ws_dash.cell(row=r, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border
ws_dash.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws_dash.cell(row=r, column=2, value="MOTIVATION STATION")

r = 22
ws_dash.cell(row=r, column=2, value="Debt-Free Countdown (Best Method)").font = body_font_bold
ws_dash.cell(row=r, column=2).border = thin_border
ws_dash.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_dash.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
for c in range(3,7):
    ws_dash.cell(row=r, column=c).border = thin_border
countdown_cell = ws_dash.cell(row=r, column=3,
    value=f'=IF({COMP}!C11=0,"Enter your debts to begin!",{COMP}!C11&" months to freedom!")')
countdown_cell.font = Font(name="Calibri", bold=True, size=13, color=TEAL)
countdown_cell.alignment = center

r = 23
ws_dash.cell(row=r, column=2, value="Monthly Payment (Min + Extra)").font = body_font_bold
ws_dash.cell(row=r, column=2).border = thin_border
ws_dash.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_dash.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
for c in range(3,7):
    ws_dash.cell(row=r, column=c).border = thin_border
ws_dash.cell(row=r, column=3,
    value=f"={DI}!E{TOTALS_ROW_INPUT}+Dashboard!{EXTRA_PMT_CELL}").number_format = CURRENCY_FMT
ws_dash.cell(row=r, column=3).font = body_font_bold
ws_dash.cell(row=r, column=3).alignment = center

r = 25
motivational_quotes = [
    '"The secret of getting ahead is getting started." -- Mark Twain',
    '"A journey of a thousand miles begins with a single step." -- Lao Tzu',
    '"Financial freedom is available to those who learn about it and work for it." -- Robert Kiyosaki',
]
for q in motivational_quotes:
    cell = ws_dash.cell(row=r, column=2, value=q)
    cell.font = Font(name="Calibri", italic=True, size=10, color=MED_GRAY)
    cell.alignment = left_wrap
    ws_dash.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1

ws_dash.freeze_panes = "B4"
ws_dash.sheet_view.showGridLines = False


# ====================================================================
# HELPER: Build a payoff plan sheet
# ====================================================================
def build_plan_sheet(ws, sheet_name, sort_method):
    ws.sheet_properties.tabColor = TEAL if sort_method == "avalanche" else GOLD

    DEBT_COL_START = 3  # C
    DEBT_COL_END = DEBT_COL_START + MAX_DEBTS - 1  # V = col 22
    TOT_PMT_COL = DEBT_COL_END + 1  # W = 23
    TOT_INT_COL = TOT_PMT_COL + 1   # X = 24
    RUN_BAL_COL = TOT_INT_COL + 1    # Y = 25

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 10
    for i in range(DEBT_COL_START, DEBT_COL_END + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions[get_column_letter(TOT_PMT_COL)].width = 16
    ws.column_dimensions[get_column_letter(TOT_INT_COL)].width = 16
    ws.column_dimensions[get_column_letter(RUN_BAL_COL)].width = 18

    last_col = RUN_BAL_COL

    title_text = "SNOWBALL PLAN  (Smallest Balance First)" if sort_method == "snowball" \
        else "AVALANCHE PLAN  (Highest Interest First)"
    title_block(ws, 1, 2, title_text, last_col, 18)
    subtitle_block(ws, 2, 2, "Month-by-month payment schedule. Green cells = debt paid off!", last_col)

    # Sorted debt list
    SORT_HDR = 4
    sort_headers = ["#", "Debt Name", "Balance", "APR %", "Min Payment"]
    sort_cols = [2, 3, 4, 5, 6]
    for ci, h in zip(sort_cols, sort_headers):
        cell = ws.cell(row=SORT_HDR, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin_border

    ws.column_dimensions["C"].width = 24

    DI_sheet = "'Debt Input'"

    # Helper columns for sorting (cols 30-35, hidden later)
    for i in range(MAX_DEBTS):
        hr = SORT_HDR + 1 + i
        di_row = DATA_START + i

        ws.cell(row=hr, column=30, value=i + 1)  # AD: index
        ws.cell(row=hr, column=31, value=f"={DI_sheet}!C{di_row}")  # AE: balance
        ws.cell(row=hr, column=32, value=f"={DI_sheet}!D{di_row}")  # AF: rate
        ws.cell(row=hr, column=33, value=f'=IF(AE{hr}>0,1,0)')  # AG: has data

        if sort_method == "snowball":
            ws.cell(row=hr, column=34,
                    value=f'=IF(AG{hr}=0,9999999999,AE{hr}+AD{hr}*0.00001)')
        else:
            ws.cell(row=hr, column=34,
                    value=f'=IF(AG{hr}=0,9999999999,-AF{hr}+AD{hr}*0.0000001)')

        ws.cell(row=hr, column=35,
                value=f'=RANK(AH{hr},AH${SORT_HDR+1}:AH${SORT_HDR+MAX_DEBTS},1)')

    SORT_DATA_START = SORT_HDR + 1
    SORT_DATA_END = SORT_HDR + MAX_DEBTS
    for pos in range(1, MAX_DEBTS + 1):
        sr = SORT_HDR + pos
        rank_range = f"AI${SORT_DATA_START}:AI${SORT_DATA_END}"

        ws.cell(row=sr, column=2, value=pos)
        ws.cell(row=sr, column=2).font = body_font
        ws.cell(row=sr, column=2).alignment = center
        ws.cell(row=sr, column=2).border = thin_border

        ws.cell(row=sr, column=3,
            value=f'=IFERROR(INDEX({DI_sheet}!B${DATA_START}:B${DATA_END},MATCH({pos},{rank_range},0)),"")').font = body_font
        ws.cell(row=sr, column=3).alignment = center
        ws.cell(row=sr, column=3).border = thin_border

        ws.cell(row=sr, column=4,
            value=f'=IFERROR(INDEX({DI_sheet}!C${DATA_START}:C${DATA_END},MATCH({pos},{rank_range},0)),0)')
        ws.cell(row=sr, column=4).number_format = CURRENCY_FMT
        ws.cell(row=sr, column=4).font = body_font
        ws.cell(row=sr, column=4).alignment = center
        ws.cell(row=sr, column=4).border = thin_border

        ws.cell(row=sr, column=5,
            value=f'=IFERROR(INDEX({DI_sheet}!D${DATA_START}:D${DATA_END},MATCH({pos},{rank_range},0)),0)')
        ws.cell(row=sr, column=5).number_format = PCT_FMT
        ws.cell(row=sr, column=5).font = body_font
        ws.cell(row=sr, column=5).alignment = center
        ws.cell(row=sr, column=5).border = thin_border

        ws.cell(row=sr, column=6,
            value=f'=IFERROR(INDEX({DI_sheet}!E${DATA_START}:E${DATA_END},MATCH({pos},{rank_range},0)),0)')
        ws.cell(row=sr, column=6).number_format = CURRENCY_FMT
        ws.cell(row=sr, column=6).font = body_font
        ws.cell(row=sr, column=6).alignment = center
        ws.cell(row=sr, column=6).border = thin_border

    apply_alt_rows(ws, SORT_DATA_START, SORT_DATA_END, 6)

    # Month-by-month schedule
    SCHED_HDR = SORT_DATA_END + 2  # row 26
    SCHED_START = SCHED_HDR + 1     # row 27

    cell = ws.cell(row=SCHED_HDR, column=2, value="Month")
    cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin_border

    for d in range(MAX_DEBTS):
        col = DEBT_COL_START + d
        sort_row = SORT_HDR + 1 + d
        cell = ws.cell(row=SCHED_HDR, column=col,
            value=f'=IF(C{sort_row}="","",C{sort_row})')
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin_border

    for col_idx, label in [(TOT_PMT_COL, "Total Payment"), (TOT_INT_COL, "Interest Paid"), (RUN_BAL_COL, "Remaining Balance")]:
        cell = ws.cell(row=SCHED_HDR, column=col_idx, value=label)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin_border

    # Helper columns per debt for schedule calculations (col 40+)
    HELPER2_START = 40

    def hcol(debt_idx, sub):
        return HELPER2_START + debt_idx * 3 + sub

    EXTRA_REF = "Dashboard!$C$11"

    print(f"Building {sheet_name} month-by-month schedule ({MAX_MONTHS} months)...")

    for month in range(1, MAX_MONTHS + 1):
        mr = SCHED_START + month - 1

        ws.cell(row=mr, column=2, value=month)
        ws.cell(row=mr, column=2).font = body_font
        ws.cell(row=mr, column=2).alignment = center
        ws.cell(row=mr, column=2).border = thin_border

        for d in range(MAX_DEBTS):
            debt_col = DEBT_COL_START + d
            debt_col_letter = get_column_letter(debt_col)
            sort_row = SORT_HDR + 1 + d

            h_interest_col = hcol(d, 0)
            h_payment_col = hcol(d, 1)
            h_extra_remain_col = hcol(d, 2)

            hi_letter = get_column_letter(h_interest_col)
            hp_letter = get_column_letter(h_payment_col)
            he_letter = get_column_letter(h_extra_remain_col)

            if month == 1:
                prev_bal = f"$D${sort_row}"
            else:
                prev_row = mr - 1
                prev_bal = f"{debt_col_letter}{prev_row}"

            rate_ref = f"$E${sort_row}"
            min_ref = f"$F${sort_row}"

            if d == 0:
                if month == 1:
                    extra_budget = f"{EXTRA_REF}"
                else:
                    prev_row = mr - 1
                    freed_parts = []
                    for dd in range(MAX_DEBTS):
                        dd_sort_row = SORT_HDR + 1 + dd
                        dd_col_letter = get_column_letter(DEBT_COL_START + dd)
                        freed_parts.append(f"IF({dd_col_letter}{prev_row}<=0,$F${dd_sort_row},0)")
                    freed_sum = "+".join(freed_parts)
                    extra_budget = f"({EXTRA_REF}+{freed_sum})"
            else:
                prev_he_letter = get_column_letter(hcol(d - 1, 2))
                extra_budget = f"{prev_he_letter}{mr}"

            # Interest
            ws.cell(row=mr, column=h_interest_col,
                value=f"=IF({prev_bal}<=0,0,{prev_bal}*{rate_ref}/12)")

            # Payment
            ws.cell(row=mr, column=h_payment_col,
                value=f"=IF({prev_bal}<=0,0,MIN({prev_bal}+{hi_letter}{mr},{min_ref}+{extra_budget}))")

            # Extra remaining
            ws.cell(row=mr, column=h_extra_remain_col,
                value=f"=IF({prev_bal}<=0,{extra_budget},MAX(0,{min_ref}+{extra_budget}-{prev_bal}-{hi_letter}{mr}))")

            # Remaining balance
            ws.cell(row=mr, column=debt_col,
                value=f"=IF({prev_bal}<=0,0,MAX(0,{prev_bal}+{hi_letter}{mr}-{hp_letter}{mr}))")
            ws.cell(row=mr, column=debt_col).number_format = CURRENCY_FMT
            ws.cell(row=mr, column=debt_col).font = body_font
            ws.cell(row=mr, column=debt_col).alignment = center
            ws.cell(row=mr, column=debt_col).border = thin_border

        # Totals for this month
        pmt_parts = [f"{get_column_letter(hcol(d, 1))}{mr}" for d in range(MAX_DEBTS)]
        ws.cell(row=mr, column=TOT_PMT_COL, value=f"={'+'.join(pmt_parts)}")
        ws.cell(row=mr, column=TOT_PMT_COL).number_format = CURRENCY_FMT
        ws.cell(row=mr, column=TOT_PMT_COL).font = body_font
        ws.cell(row=mr, column=TOT_PMT_COL).alignment = center
        ws.cell(row=mr, column=TOT_PMT_COL).border = thin_border

        int_parts = [f"{get_column_letter(hcol(d, 0))}{mr}" for d in range(MAX_DEBTS)]
        ws.cell(row=mr, column=TOT_INT_COL, value=f"={'+'.join(int_parts)}")
        ws.cell(row=mr, column=TOT_INT_COL).number_format = CURRENCY_FMT
        ws.cell(row=mr, column=TOT_INT_COL).font = body_font
        ws.cell(row=mr, column=TOT_INT_COL).alignment = center
        ws.cell(row=mr, column=TOT_INT_COL).border = thin_border

        bal_parts = [f"{get_column_letter(DEBT_COL_START + d)}{mr}" for d in range(MAX_DEBTS)]
        ws.cell(row=mr, column=RUN_BAL_COL, value=f"={'+'.join(bal_parts)}")
        ws.cell(row=mr, column=RUN_BAL_COL).number_format = CURRENCY_FMT
        ws.cell(row=mr, column=RUN_BAL_COL).font = body_font_bold
        ws.cell(row=mr, column=RUN_BAL_COL).alignment = center
        ws.cell(row=mr, column=RUN_BAL_COL).border = thin_border

    # Conditional formatting: green when paid off
    sched_end_row = SCHED_START + MAX_MONTHS - 1
    for d in range(MAX_DEBTS):
        col_letter = get_column_letter(DEBT_COL_START + d)
        cell_range = f"{col_letter}{SCHED_START}:{col_letter}{sched_end_row}"
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f"AND({col_letter}{SCHED_START}=0,$D${SORT_HDR+1+d}>0)"],
                fill=PatternFill("solid", fgColor=TEAL_LIGHT),
            )
        )

    ws.freeze_panes = f"C{SCHED_START}"
    ws.sheet_view.showGridLines = False

    return {
        "total_interest_col": get_column_letter(TOT_INT_COL),
        "total_payment_col": get_column_letter(TOT_PMT_COL),
        "running_bal_col": get_column_letter(RUN_BAL_COL),
        "sched_start": SCHED_START,
        "sched_end": sched_end_row,
        "sheet_name": sheet_name,
    }


# ====================================================================
# SHEET 4 - SNOWBALL PLAN
# ====================================================================
ws_snow = wb.create_sheet("Snowball Plan")
snow_info = build_plan_sheet(ws_snow, "Snowball Plan", "snowball")

# ====================================================================
# SHEET 5 - AVALANCHE PLAN
# ====================================================================
ws_aval = wb.create_sheet("Avalanche Plan")
aval_info = build_plan_sheet(ws_aval, "Avalanche Plan", "avalanche")

# ====================================================================
# SHEET 6 - COMPARISON
# ====================================================================
ws_comp = wb.create_sheet("Comparison")
ws_comp.sheet_properties.tabColor = GOLD

set_col_widths(ws_comp, {"A": 4, "B": 36, "C": 22, "D": 22, "E": 4})

title_block(ws_comp, 1, 2, "METHOD COMPARISON", 4, 20)
subtitle_block(ws_comp, 2, 2, "Snowball vs. Avalanche -- side by side", 4)

r = 4
for ci, txt in [(2, "Metric"), (3, "Snowball"), (4, "Avalanche")]:
    cell = ws_comp.cell(row=r, column=ci, value=txt)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin_border

SNOW = "'Snowball Plan'"
AVAL = "'Avalanche Plan'"

snow_bal_range = f"{SNOW}!{snow_info['running_bal_col']}{snow_info['sched_start']}:{snow_info['running_bal_col']}{snow_info['sched_end']}"
aval_bal_range = f"{AVAL}!{aval_info['running_bal_col']}{aval_info['sched_start']}:{aval_info['running_bal_col']}{aval_info['sched_end']}"
snow_int_range = f"{SNOW}!{snow_info['total_interest_col']}{snow_info['sched_start']}:{snow_info['total_interest_col']}{snow_info['sched_end']}"
aval_int_range = f"{AVAL}!{aval_info['total_interest_col']}{aval_info['sched_start']}:{aval_info['total_interest_col']}{aval_info['sched_end']}"
snow_pmt_range = f"{SNOW}!{snow_info['total_payment_col']}{snow_info['sched_start']}:{snow_info['total_payment_col']}{snow_info['sched_end']}"
aval_pmt_range = f"{AVAL}!{aval_info['total_payment_col']}{aval_info['sched_start']}:{aval_info['total_payment_col']}{aval_info['sched_end']}"

# Row 5: Total months
r = 5
ws_comp.cell(row=r, column=2, value="Total Months to Payoff").font = body_font_bold
ws_comp.cell(row=r, column=2).border = thin_border
ws_comp.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_comp.cell(row=r, column=3,
    value=f"=IFERROR(MATCH(TRUE,INDEX({snow_bal_range}<=0,0),0),120)").number_format = INT_FMT
ws_comp.cell(row=r, column=3).font = body_font; ws_comp.cell(row=r, column=3).alignment = center; ws_comp.cell(row=r, column=3).border = thin_border
ws_comp.cell(row=r, column=4,
    value=f"=IFERROR(MATCH(TRUE,INDEX({aval_bal_range}<=0,0),0),120)").number_format = INT_FMT
ws_comp.cell(row=r, column=4).font = body_font; ws_comp.cell(row=r, column=4).alignment = center; ws_comp.cell(row=r, column=4).border = thin_border

# Row 6: Total interest
r = 6
ws_comp.cell(row=r, column=2, value="Total Interest Paid").font = body_font_bold
ws_comp.cell(row=r, column=2).border = thin_border
ws_comp.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_comp.cell(row=r, column=3, value=f"=SUM({snow_int_range})").number_format = CURRENCY_FMT
ws_comp.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=11, color=RED)
ws_comp.cell(row=r, column=3).alignment = center; ws_comp.cell(row=r, column=3).border = thin_border
ws_comp.cell(row=r, column=4, value=f"=SUM({aval_int_range})").number_format = CURRENCY_FMT
ws_comp.cell(row=r, column=4).font = Font(name="Calibri", bold=True, size=11, color=RED)
ws_comp.cell(row=r, column=4).alignment = center; ws_comp.cell(row=r, column=4).border = thin_border

# Row 7: Total amount paid
r = 7
ws_comp.cell(row=r, column=2, value="Total Amount Paid").font = body_font_bold
ws_comp.cell(row=r, column=2).border = thin_border
ws_comp.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_comp.cell(row=r, column=3, value=f"=SUM({snow_pmt_range})").number_format = CURRENCY_FMT
ws_comp.cell(row=r, column=3).font = body_font; ws_comp.cell(row=r, column=3).alignment = center; ws_comp.cell(row=r, column=3).border = thin_border
ws_comp.cell(row=r, column=4, value=f"=SUM({aval_pmt_range})").number_format = CURRENCY_FMT
ws_comp.cell(row=r, column=4).font = body_font; ws_comp.cell(row=r, column=4).alignment = center; ws_comp.cell(row=r, column=4).border = thin_border

# Row 9: Interest saved
r = 9
ws_comp.cell(row=r, column=2, value="Interest Saved (Best Method)").font = Font(name="Calibri", bold=True, size=12, color=TEAL)
ws_comp.cell(row=r, column=2).border = thin_border
ws_comp.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_comp.merge_cells("C9:D9")
ws_comp.cell(row=r, column=3, value="=ABS(C6-D6)").number_format = CURRENCY_FMT
ws_comp.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=14, color=TEAL)
ws_comp.cell(row=r, column=3).fill = teal_light_fill; ws_comp.cell(row=r, column=3).alignment = center; ws_comp.cell(row=r, column=3).border = thin_border

# Row 10: Recommended
r = 10
ws_comp.cell(row=r, column=2, value="Recommended Method").font = Font(name="Calibri", bold=True, size=12, color=CHARCOAL)
ws_comp.cell(row=r, column=2).border = thin_border
ws_comp.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_comp.merge_cells("C10:D10")
ws_comp.cell(row=r, column=3,
    value='=IF(C6<D6,"Snowball saves more!",IF(D6<C6,"Avalanche saves more!","Tied -- both are equal!"))').font = Font(name="Calibri", bold=True, size=14, color=CHARCOAL)
ws_comp.cell(row=r, column=3).fill = gold_light_fill; ws_comp.cell(row=r, column=3).alignment = center; ws_comp.cell(row=r, column=3).border = thin_border

# Row 11: Best months
r = 11
ws_comp.cell(row=r, column=2, value="Best Months to Payoff").font = body_font_bold
ws_comp.cell(row=r, column=2).border = thin_border
ws_comp.merge_cells("C11:D11")
ws_comp.cell(row=r, column=3, value="=MIN(C5,D5)").number_format = INT_FMT
ws_comp.cell(row=r, column=3).font = body_font_bold; ws_comp.cell(row=r, column=3).alignment = center; ws_comp.cell(row=r, column=3).border = thin_border

# Conditional formatting: winner in green
for row_ref in ["5", "6", "7"]:
    ws_comp.conditional_formatting.add(
        f"C{row_ref}",
        CellIsRule(operator="lessThan", formula=[f"D{row_ref}"],
                   fill=PatternFill("solid", fgColor=TEAL_LIGHT),
                   font=Font(bold=True, color=TEAL))
    )
    ws_comp.conditional_formatting.add(
        f"D{row_ref}",
        CellIsRule(operator="lessThan", formula=[f"C{row_ref}"],
                   fill=PatternFill("solid", fgColor=TEAL_LIGHT),
                   font=Font(bold=True, color=TEAL))
    )

# Detailed breakdown
r = 13
for c in range(2, 5):
    cell = ws_comp.cell(row=r, column=c)
    cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = thin_border
ws_comp.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
ws_comp.cell(row=r, column=2, value="DETAILED BREAKDOWN")

r = 14
for ci, txt in [(2, "Detail"), (3, "Snowball"), (4, "Avalanche")]:
    cell = ws_comp.cell(row=r, column=ci, value=txt)
    cell.font = sub_header_font; cell.fill = sub_header_fill; cell.alignment = center; cell.border = thin_border

r = 15
ws_comp.cell(row=r, column=2, value="Months Difference").font = body_font_bold
ws_comp.cell(row=r, column=2).border = thin_border
ws_comp.merge_cells("C15:D15")
ws_comp.cell(row=r, column=3, value="=ABS(C5-D5)").number_format = INT_FMT
ws_comp.cell(row=r, column=3).font = body_font; ws_comp.cell(row=r, column=3).alignment = center; ws_comp.cell(row=r, column=3).border = thin_border

r = 16
ws_comp.cell(row=r, column=2, value="Interest Difference").font = body_font_bold
ws_comp.cell(row=r, column=2).border = thin_border
ws_comp.merge_cells("C16:D16")
ws_comp.cell(row=r, column=3, value="=ABS(C6-D6)").number_format = CURRENCY_FMT
ws_comp.cell(row=r, column=3).font = body_font; ws_comp.cell(row=r, column=3).alignment = center; ws_comp.cell(row=r, column=3).border = thin_border

r = 17
ws_comp.cell(row=r, column=2, value="Total Paid Difference").font = body_font_bold
ws_comp.cell(row=r, column=2).border = thin_border
ws_comp.merge_cells("C17:D17")
ws_comp.cell(row=r, column=3, value="=ABS(C7-D7)").number_format = CURRENCY_FMT
ws_comp.cell(row=r, column=3).font = body_font; ws_comp.cell(row=r, column=3).alignment = center; ws_comp.cell(row=r, column=3).border = thin_border

r = 19
note_text = (
    "NOTE: The Avalanche method typically saves the most on interest. "
    "The Snowball method pays off individual debts faster for psychological wins. "
    "Choose the method that keeps YOU motivated!"
)
ws_comp.merge_cells("B19:D21")
cell = ws_comp.cell(row=19, column=2, value=note_text)
cell.font = Font(name="Calibri", italic=True, size=11, color=MED_GRAY)
cell.alignment = left_wrap

ws_comp.freeze_panes = "B5"
ws_comp.sheet_view.showGridLines = False


# ====================================================================
# FINAL: Hide helper columns, save
# ====================================================================
for ws_plan in [ws_snow, ws_aval]:
    for c in range(30, 36):
        ws_plan.column_dimensions[get_column_letter(c)].hidden = True
    for d in range(MAX_DEBTS):
        for sub in range(3):
            col = 40 + d * 3 + sub
            if col <= 16384:  # Excel max columns
                ws_plan.column_dimensions[get_column_letter(col)].hidden = True

wb.active = wb.sheetnames.index("Dashboard")

OUTPUT = "/Users/timdunn/mobile_app_ideas/etsy-templates/debt-payoff-calculator.xlsx"
wb.save(OUTPUT)
print(f"SUCCESS: Created {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"File saved successfully!")
