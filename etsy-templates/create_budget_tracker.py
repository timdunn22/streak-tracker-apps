#!/usr/bin/env python3
"""
Monthly Budget Tracker - Premium Etsy Template
Creates a professional Excel workbook with 5 sheets.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers, Protection, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from copy import copy
import datetime

# ─── Color Palette ──────────────────────────────────────────────
NAVY        = "1B2A4A"
NAVY_LIGHT  = "2D4A7A"
WHITE       = "FFFFFF"
LIGHT_BLUE  = "E8F0FE"
MID_BLUE    = "C5D9F1"
EMERALD     = "10B981"
EMERALD_BG  = "D1FAE5"
RED         = "EF4444"
RED_BG      = "FEE2E2"
AMBER       = "F59E0B"
AMBER_BG    = "FEF3C7"
GRAY_LIGHT  = "F3F4F6"
GRAY_MED    = "D1D5DB"
GRAY_DARK   = "6B7280"
BLACK       = "000000"

# ─── Reusable Styles ───────────────────────────────────────────
thin_border = Border(
    left=Side(style='thin', color=GRAY_MED),
    right=Side(style='thin', color=GRAY_MED),
    top=Side(style='thin', color=GRAY_MED),
    bottom=Side(style='thin', color=GRAY_MED),
)
medium_border = Border(
    left=Side(style='medium', color=NAVY),
    right=Side(style='medium', color=NAVY),
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='medium', color=NAVY),
)
header_border = Border(
    left=Side(style='thin', color=NAVY_LIGHT),
    right=Side(style='thin', color=NAVY_LIGHT),
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='medium', color=NAVY),
)
bottom_thick = Border(
    left=Side(style='thin', color=GRAY_MED),
    right=Side(style='thin', color=GRAY_MED),
    top=Side(style='thin', color=GRAY_MED),
    bottom=Side(style='medium', color=NAVY),
)

navy_fill    = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
light_fill   = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
mid_fill     = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type='solid')
white_fill   = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
emerald_fill = PatternFill(start_color=EMERALD_BG, end_color=EMERALD_BG, fill_type='solid')
red_fill_bg  = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type='solid')
amber_fill   = PatternFill(start_color=AMBER_BG, end_color=AMBER_BG, fill_type='solid')
gray_fill    = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type='solid')
emerald_solid= PatternFill(start_color=EMERALD, end_color=EMERALD, fill_type='solid')
red_solid    = PatternFill(start_color=RED, end_color=RED, fill_type='solid')

font_title    = Font(name='Calibri', size=22, bold=True, color=NAVY)
font_subtitle = Font(name='Calibri', size=14, bold=True, color=NAVY)
font_header   = Font(name='Calibri', size=11, bold=True, color=WHITE)
font_header_dark = Font(name='Calibri', size=11, bold=True, color=NAVY)
font_body     = Font(name='Calibri', size=11, color=BLACK)
font_body_gray= Font(name='Calibri', size=11, color=GRAY_DARK)
font_small    = Font(name='Calibri', size=9, color=GRAY_DARK)
font_money    = Font(name='Calibri', size=11, color=BLACK)
font_money_bold = Font(name='Calibri', size=11, bold=True, color=NAVY)
font_positive = Font(name='Calibri', size=11, bold=True, color=EMERALD)
font_negative = Font(name='Calibri', size=11, bold=True, color=RED)
font_white    = Font(name='Calibri', size=11, color=WHITE)
font_white_bold = Font(name='Calibri', size=12, bold=True, color=WHITE)
font_link     = Font(name='Calibri', size=11, color="2563EB", underline='single')

align_center  = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left    = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right   = Alignment(horizontal='right', vertical='center')
align_top     = Alignment(horizontal='left', vertical='top', wrap_text=True)

CURRENCY_FMT  = '"$"#,##0.00'
PERCENT_FMT   = '0.0%'
DATE_FMT      = 'MM/DD/YYYY'

CATEGORIES = [
    "Housing", "Transportation", "Food & Groceries", "Utilities",
    "Insurance", "Healthcare", "Debt Payments", "Personal",
    "Entertainment", "Savings", "Education", "Miscellaneous"
]

PAYMENT_METHODS = ["Cash", "Credit Card", "Debit Card", "Bank Transfer", "Venmo/Zelle", "Other"]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

SAMPLE_BUDGETS = {
    "Housing": 1500, "Transportation": 400, "Food & Groceries": 600,
    "Utilities": 250, "Insurance": 300, "Healthcare": 150,
    "Debt Payments": 500, "Personal": 200, "Entertainment": 150,
    "Savings": 500, "Education": 100, "Miscellaneous": 100,
}


def set_print_settings(ws):
    """Apply print-friendly settings."""
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


def style_range_as_header(ws, row, col_start, col_end, fill=None, font=None, border=None):
    """Apply header styling to a range of cells in a single row."""
    _fill = fill or navy_fill
    _font = font or font_header
    _border = border or header_border
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill
        cell.font = _font
        cell.border = _border
        cell.alignment = align_center


def make_section_header(ws, row, col_start, col_end, text, fill=None):
    """Create a merged section header."""
    _fill = fill or navy_fill
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.fill = _fill
    cell.font = font_white_bold
    cell.alignment = align_center
    cell.border = medium_border
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=c).fill = _fill
        ws.cell(row=row, column=c).border = medium_border


# ═══════════════════════════════════════════════════════════════
# CREATE WORKBOOK
# ═══════════════════════════════════════════════════════════════
wb = Workbook()

# ───────────────────────────────────────────────────────────────
# SHEET 1: INSTRUCTIONS
# ───────────────────────────────────────────────────────────────
ws_inst = wb.active
ws_inst.title = "Instructions"
ws_inst.sheet_properties.tabColor = NAVY

# Column widths
ws_inst.column_dimensions['A'].width = 3
ws_inst.column_dimensions['B'].width = 5
ws_inst.column_dimensions['C'].width = 70
ws_inst.column_dimensions['D'].width = 3

set_print_settings(ws_inst)

# Background fill for the whole visible area
for r in range(1, 65):
    for c in range(1, 5):
        ws_inst.cell(row=r, column=c).fill = white_fill

# ── Title Banner ──
ws_inst.merge_cells('A1:D3')
title_cell = ws_inst['A1']
title_cell.value = "MONTHLY BUDGET TRACKER"
title_cell.font = Font(name='Calibri', size=28, bold=True, color=WHITE)
title_cell.fill = navy_fill
title_cell.alignment = Alignment(horizontal='center', vertical='center')
for r in range(1, 4):
    for c in range(1, 5):
        ws_inst.cell(row=r, column=c).fill = navy_fill
        ws_inst.cell(row=r, column=c).border = Border()

# Subtitle
ws_inst.merge_cells('A4:D4')
sub = ws_inst['A4']
sub.value = "Your Complete Personal Finance Companion"
sub.font = Font(name='Calibri', size=13, italic=True, color=NAVY_LIGHT)
sub.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
sub.alignment = Alignment(horizontal='center', vertical='center')
ws_inst.row_dimensions[4].height = 30
for c in range(1, 5):
    ws_inst.cell(row=4, column=c).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')

row = 6

# ── Welcome ──
ws_inst.merge_cells(f'B{row}:C{row}')
ws_inst.cell(row=row, column=2).value = "Welcome!"
ws_inst.cell(row=row, column=2).font = Font(name='Calibri', size=16, bold=True, color=NAVY)
row += 1

ws_inst.merge_cells(f'B{row}:C{row}')
ws_inst.row_dimensions[row].height = 50
ws_inst.cell(row=row, column=2).value = (
    "Thank you for purchasing this budget tracker! This workbook is designed to help you "
    "take control of your finances with a clear, easy-to-follow system. Simply follow the "
    "steps below to get started."
)
ws_inst.cell(row=row, column=2).font = font_body
ws_inst.cell(row=row, column=2).alignment = align_top
row += 2

# ── How To Use ──
ws_inst.merge_cells(f'B{row}:C{row}')
ws_inst.cell(row=row, column=2).value = "How To Use This Workbook"
ws_inst.cell(row=row, column=2).font = Font(name='Calibri', size=14, bold=True, color=NAVY)
ws_inst.cell(row=row, column=2).border = Border(bottom=Side(style='medium', color=NAVY))
ws_inst.cell(row=row, column=3).border = Border(bottom=Side(style='medium', color=NAVY))
row += 2

steps = [
    ("Step 1: Set Your Budget", "Monthly Budget",
     "Go to the 'Monthly Budget' sheet and enter your budgeted amounts for each category. "
     "These are your spending targets for the month."),
    ("Step 2: Log Transactions", "Transactions",
     "Every time you spend or receive money, log it in the 'Transactions' sheet. "
     "Use the dropdown menus for Category and Payment Method to keep entries consistent."),
    ("Step 3: Review Your Dashboard", "Dashboard",
     "The 'Dashboard' sheet automatically calculates your totals, shows your spending "
     "vs. budget by category, and highlights areas where you're over or under budget."),
    ("Step 4: Track Annual Progress", "Annual Overview",
     "At the end of each month, the 'Annual Overview' sheet gives you a bird's-eye view "
     "of your spending patterns across the entire year."),
]

for title, sheet_name, desc in steps:
    # Step title row with colored left bar
    ws_inst.cell(row=row, column=2).value = ">"
    ws_inst.cell(row=row, column=2).font = Font(name='Calibri', size=11, bold=True, color=EMERALD)
    ws_inst.cell(row=row, column=2).fill = emerald_fill
    ws_inst.cell(row=row, column=2).alignment = align_center
    ws_inst.merge_cells(f'C{row}:C{row}')
    ws_inst.cell(row=row, column=3).value = f"{title}  (Sheet: \"{sheet_name}\")"
    ws_inst.cell(row=row, column=3).font = Font(name='Calibri', size=11, bold=True, color=NAVY)
    ws_inst.cell(row=row, column=3).fill = emerald_fill
    row += 1
    # Description
    ws_inst.merge_cells(f'B{row}:C{row}')
    ws_inst.cell(row=row, column=2).value = desc
    ws_inst.cell(row=row, column=2).font = font_body
    ws_inst.cell(row=row, column=2).alignment = align_top
    ws_inst.row_dimensions[row].height = 40
    row += 2

# ── Color Legend ──
ws_inst.merge_cells(f'B{row}:C{row}')
ws_inst.cell(row=row, column=2).value = "Color-Coded Legend"
ws_inst.cell(row=row, column=2).font = Font(name='Calibri', size=14, bold=True, color=NAVY)
ws_inst.cell(row=row, column=2).border = Border(bottom=Side(style='medium', color=NAVY))
ws_inst.cell(row=row, column=3).border = Border(bottom=Side(style='medium', color=NAVY))
row += 2

legend_items = [
    (emerald_fill, Font(name='Calibri', size=11, bold=True, color=EMERALD), "Green", "Under budget / Positive cash flow"),
    (red_fill_bg, Font(name='Calibri', size=11, bold=True, color=RED), "Red", "Over budget / Negative cash flow"),
    (amber_fill, Font(name='Calibri', size=11, bold=True, color=AMBER), "Amber", "Approaching budget limit (within 10%)"),
    (navy_fill, font_white_bold, "Navy", "Section headers and titles"),
    (light_fill, font_body, "Light Blue", "Alternating row shading for readability"),
]

for fill_color, label_font, label, meaning in legend_items:
    ws_inst.cell(row=row, column=2).fill = fill_color
    ws_inst.cell(row=row, column=2).value = f"  {label}"
    ws_inst.cell(row=row, column=2).font = label_font
    ws_inst.cell(row=row, column=2).alignment = align_center
    ws_inst.cell(row=row, column=2).border = thin_border
    ws_inst.cell(row=row, column=3).value = meaning
    ws_inst.cell(row=row, column=3).font = font_body
    ws_inst.cell(row=row, column=3).border = thin_border
    ws_inst.row_dimensions[row].height = 25
    row += 1

row += 1

# ── Tips ──
ws_inst.merge_cells(f'B{row}:C{row}')
ws_inst.cell(row=row, column=2).value = "Budgeting Tips"
ws_inst.cell(row=row, column=2).font = Font(name='Calibri', size=14, bold=True, color=NAVY)
ws_inst.cell(row=row, column=2).border = Border(bottom=Side(style='medium', color=NAVY))
ws_inst.cell(row=row, column=3).border = Border(bottom=Side(style='medium', color=NAVY))
row += 2

tips = [
    "Follow the 50/30/20 rule: 50% needs, 30% wants, 20% savings & debt repayment.",
    "Log transactions daily -- even small purchases add up over a month.",
    "Review your Dashboard weekly to catch overspending before it becomes a problem.",
    "Build an emergency fund of 3-6 months of expenses before aggressive investing.",
    "Use the 24-hour rule: wait a day before any non-essential purchase over $50.",
    "Automate your savings -- treat it like a bill that must be paid each month.",
    "Adjust your budget monthly. Life changes, and your budget should too.",
]

for i, tip in enumerate(tips, 1):
    ws_inst.cell(row=row, column=2).value = str(i)
    ws_inst.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    ws_inst.cell(row=row, column=2).fill = navy_fill
    ws_inst.cell(row=row, column=2).alignment = align_center
    ws_inst.cell(row=row, column=2).border = thin_border
    ws_inst.cell(row=row, column=3).value = tip
    ws_inst.cell(row=row, column=3).font = font_body
    ws_inst.cell(row=row, column=3).alignment = align_left
    ws_inst.cell(row=row, column=3).border = thin_border
    ws_inst.row_dimensions[row].height = 28
    row += 1

row += 2
ws_inst.merge_cells(f'B{row}:C{row}')
ws_inst.cell(row=row, column=2).value = "Do not modify cells with formulas (gray text). Only edit cells with white backgrounds."
ws_inst.cell(row=row, column=2).font = Font(name='Calibri', size=10, italic=True, color=RED)
ws_inst.cell(row=row, column=2).alignment = align_center

# Freeze
ws_inst.freeze_panes = 'A5'


# ───────────────────────────────────────────────────────────────
# SHEET 2: DASHBOARD
# ───────────────────────────────────────────────────────────────
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_properties.tabColor = EMERALD

set_print_settings(ws_dash)

# Column widths
col_widths_dash = {'A': 3, 'B': 22, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 16, 'H': 3}
for col, w in col_widths_dash.items():
    ws_dash.column_dimensions[col].width = w

# ── Title Banner ──
ws_dash.merge_cells('A1:H2')
ws_dash['A1'].value = "FINANCIAL DASHBOARD"
ws_dash['A1'].font = Font(name='Calibri', size=22, bold=True, color=WHITE)
ws_dash['A1'].fill = navy_fill
ws_dash['A1'].alignment = align_center
for r in range(1, 3):
    for c in range(1, 9):
        ws_dash.cell(row=r, column=c).fill = navy_fill

# ── Summary KPI Cards ──
row = 4
make_section_header(ws_dash, row, 2, 7, "MONTHLY SUMMARY")
row = 5

kpi_labels = ["Total Income", "Total Expenses", "Net Savings", "Savings Rate", "Largest Expense Category", "Days Left in Month"]
kpi_cols = [(2, 3), (4, 5), (6, 7)]

# Row 5: labels for first 3 KPIs
for i, (cs, ce) in enumerate(kpi_cols):
    ws_dash.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws_dash.cell(row=row, column=cs)
    cell.value = kpi_labels[i]
    cell.font = Font(name='Calibri', size=10, bold=True, color=GRAY_DARK)
    cell.fill = light_fill
    cell.alignment = align_center
    cell.border = thin_border
    ws_dash.cell(row=row, column=ce).border = thin_border
    ws_dash.cell(row=row, column=ce).fill = light_fill

ws_dash.row_dimensions[5].height = 22

# Row 6: values for first 3 KPIs
row = 6
ws_dash.row_dimensions[6].height = 36

# Total Income: sum positive amounts from Transactions
for i, (cs, ce) in enumerate(kpi_cols):
    ws_dash.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws_dash.cell(row=row, column=cs)
    cell.border = medium_border
    ws_dash.cell(row=row, column=ce).border = medium_border

# Total Income
ws_dash.cell(row=6, column=2).value = "=SUMPRODUCT((Transactions!D2:D501>0)*Transactions!D2:D501)"
ws_dash.cell(row=6, column=2).number_format = CURRENCY_FMT
ws_dash.cell(row=6, column=2).font = Font(name='Calibri', size=18, bold=True, color=EMERALD)
ws_dash.cell(row=6, column=2).alignment = align_center

# Total Expenses
ws_dash.cell(row=6, column=4).value = "=SUMPRODUCT((Transactions!D2:D501<0)*Transactions!D2:D501)*-1"
ws_dash.cell(row=6, column=4).number_format = CURRENCY_FMT
ws_dash.cell(row=6, column=4).font = Font(name='Calibri', size=18, bold=True, color=RED)
ws_dash.cell(row=6, column=4).alignment = align_center

# Net Savings
ws_dash.cell(row=6, column=6).value = "=B6-D6"
ws_dash.cell(row=6, column=6).number_format = CURRENCY_FMT
ws_dash.cell(row=6, column=6).font = Font(name='Calibri', size=18, bold=True, color=NAVY)
ws_dash.cell(row=6, column=6).alignment = align_center

# Row 7-8: second row of KPIs
row = 8
for i, (cs, ce) in enumerate(kpi_cols):
    ws_dash.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws_dash.cell(row=row, column=cs)
    cell.value = kpi_labels[i + 3]
    cell.font = Font(name='Calibri', size=10, bold=True, color=GRAY_DARK)
    cell.fill = light_fill
    cell.alignment = align_center
    cell.border = thin_border
    ws_dash.cell(row=row, column=ce).fill = light_fill
    ws_dash.cell(row=row, column=ce).border = thin_border

row = 9
ws_dash.row_dimensions[9].height = 36
for i, (cs, ce) in enumerate(kpi_cols):
    ws_dash.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws_dash.cell(row=row, column=cs)
    cell.border = medium_border
    ws_dash.cell(row=row, column=ce).border = medium_border

# Savings Rate
ws_dash.cell(row=9, column=2).value = '=IF(B6=0,0,F6/B6)'
ws_dash.cell(row=9, column=2).number_format = PERCENT_FMT
ws_dash.cell(row=9, column=2).font = Font(name='Calibri', size=18, bold=True, color=NAVY)
ws_dash.cell(row=9, column=2).alignment = align_center

# Largest Expense Category
ws_dash.cell(row=9, column=4).value = '=IFERROR(INDEX(\'Monthly Budget\'!B5:B16,MATCH(MAX(\'Monthly Budget\'!D5:D16),\'Monthly Budget\'!D5:D16,0)),"-")'
ws_dash.cell(row=9, column=4).font = Font(name='Calibri', size=14, bold=True, color=NAVY)
ws_dash.cell(row=9, column=4).alignment = align_center

# Placeholder for days remaining
ws_dash.cell(row=9, column=6).value = '=DAY(EOMONTH(TODAY(),0))-DAY(TODAY())'
ws_dash.cell(row=9, column=6).font = Font(name='Calibri', size=18, bold=True, color=NAVY)
ws_dash.cell(row=9, column=6).alignment = align_center

# Conditional formatting on net savings
green_font = Font(color=EMERALD)
red_font = Font(color=RED)
ws_dash.conditional_formatting.add('F6',
    CellIsRule(operator='greaterThanOrEqual', formula=['0'],
               fill=emerald_fill, font=Font(name='Calibri', size=18, bold=True, color=EMERALD)))
ws_dash.conditional_formatting.add('F6',
    CellIsRule(operator='lessThan', formula=['0'],
               fill=red_fill_bg, font=Font(name='Calibri', size=18, bold=True, color=RED)))

# ── Category Breakdown ──
row = 11
make_section_header(ws_dash, row, 2, 7, "BUDGET vs. ACTUAL BY CATEGORY")
row = 12
cat_headers = ["Category", "Budgeted", "Actual Spent", "Difference", "% Used", "Status"]
for i, h in enumerate(cat_headers):
    cell = ws_dash.cell(row=row, column=i + 2)
    cell.value = h
    cell.font = font_header
    cell.fill = navy_fill
    cell.alignment = align_center
    cell.border = header_border

for idx, cat in enumerate(CATEGORIES):
    r = row + 1 + idx
    bg = light_fill if idx % 2 == 0 else white_fill
    bref = f"'Monthly Budget'!C{5 + idx}"
    aref = f"'Monthly Budget'!D{5 + idx}"

    ws_dash.cell(row=r, column=2, value=cat).font = Font(name='Calibri', size=11, bold=True, color=NAVY)
    ws_dash.cell(row=r, column=2).fill = bg
    ws_dash.cell(row=r, column=2).border = thin_border

    # Budgeted
    ws_dash.cell(row=r, column=3).value = f"={bref}"
    ws_dash.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws_dash.cell(row=r, column=3).font = font_money
    ws_dash.cell(row=r, column=3).fill = bg
    ws_dash.cell(row=r, column=3).border = thin_border
    ws_dash.cell(row=r, column=3).alignment = align_right

    # Actual
    ws_dash.cell(row=r, column=4).value = f"={aref}"
    ws_dash.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws_dash.cell(row=r, column=4).font = font_money
    ws_dash.cell(row=r, column=4).fill = bg
    ws_dash.cell(row=r, column=4).border = thin_border
    ws_dash.cell(row=r, column=4).alignment = align_right

    # Difference
    ws_dash.cell(row=r, column=5).value = f"=C{r}-D{r}"
    ws_dash.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws_dash.cell(row=r, column=5).font = font_money
    ws_dash.cell(row=r, column=5).fill = bg
    ws_dash.cell(row=r, column=5).border = thin_border
    ws_dash.cell(row=r, column=5).alignment = align_right

    # % Used
    ws_dash.cell(row=r, column=6).value = f'=IF(C{r}=0,0,D{r}/C{r})'
    ws_dash.cell(row=r, column=6).number_format = PERCENT_FMT
    ws_dash.cell(row=r, column=6).font = font_money
    ws_dash.cell(row=r, column=6).fill = bg
    ws_dash.cell(row=r, column=6).border = thin_border
    ws_dash.cell(row=r, column=6).alignment = align_center

    # Status
    ws_dash.cell(row=r, column=7).value = f'=IF(D{r}=0,"-",IF(D{r}<=C{r}*0.9,"Under Budget",IF(D{r}<=C{r},"Near Limit","OVER BUDGET")))'
    ws_dash.cell(row=r, column=7).font = font_body
    ws_dash.cell(row=r, column=7).fill = bg
    ws_dash.cell(row=r, column=7).border = thin_border
    ws_dash.cell(row=r, column=7).alignment = align_center

# Conditional formatting for Difference column and Status
diff_range = f"E13:E24"
ws_dash.conditional_formatting.add(diff_range,
    CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=emerald_fill, font=font_positive))
ws_dash.conditional_formatting.add(diff_range,
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill_bg, font=font_negative))

status_range = f"G13:G24"
ws_dash.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"Under Budget"'], fill=emerald_fill, font=font_positive))
ws_dash.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"OVER BUDGET"'], fill=red_fill_bg, font=font_negative))
ws_dash.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"Near Limit"'], fill=amber_fill,
               font=Font(name='Calibri', size=11, bold=True, color=AMBER)))

# Totals row
total_row = 25
ws_dash.cell(row=total_row, column=2, value="TOTAL").font = font_white_bold
ws_dash.cell(row=total_row, column=2).fill = navy_fill
ws_dash.cell(row=total_row, column=2).border = medium_border
ws_dash.cell(row=total_row, column=2).alignment = align_center
for c in range(3, 8):
    ws_dash.cell(row=total_row, column=c).fill = navy_fill
    ws_dash.cell(row=total_row, column=c).border = medium_border
    ws_dash.cell(row=total_row, column=c).font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    ws_dash.cell(row=total_row, column=c).alignment = align_center

ws_dash.cell(row=total_row, column=3).value = "=SUM(C13:C24)"
ws_dash.cell(row=total_row, column=3).number_format = CURRENCY_FMT
ws_dash.cell(row=total_row, column=4).value = "=SUM(D13:D24)"
ws_dash.cell(row=total_row, column=4).number_format = CURRENCY_FMT
ws_dash.cell(row=total_row, column=5).value = "=C25-D25"
ws_dash.cell(row=total_row, column=5).number_format = CURRENCY_FMT
ws_dash.cell(row=total_row, column=6).value = '=IF(C25=0,0,D25/C25)'
ws_dash.cell(row=total_row, column=6).number_format = PERCENT_FMT

# ── Monthly Comparison ──
row = 27
make_section_header(ws_dash, row, 2, 7, "MONTHLY INCOME & EXPENSES COMPARISON")
row = 28

# We'll put month labels in row 28, income in 29, expenses in 30, net in 31
headers_mc = ["", "Income", "Expenses", "Net", "Savings Rate"]
# Actually, let's do a transposed layout: months down the left
mc_headers = ["Month", "Income", "Expenses", "Net Savings", "Savings Rate"]
for i, h in enumerate(mc_headers):
    cell = ws_dash.cell(row=row, column=i + 2)
    cell.value = h
    cell.font = font_header
    cell.fill = navy_fill
    cell.alignment = align_center
    cell.border = header_border

for mi, month in enumerate(MONTHS):
    r = row + 1 + mi
    bg = light_fill if mi % 2 == 0 else white_fill
    ws_dash.cell(row=r, column=2, value=month).font = Font(name='Calibri', size=11, bold=True, color=NAVY)
    ws_dash.cell(row=r, column=2).fill = bg
    ws_dash.cell(row=r, column=2).border = thin_border
    ws_dash.cell(row=r, column=2).alignment = align_center

    # Reference Annual Overview sheet
    ao_col = get_column_letter(mi + 3)  # C=Jan, D=Feb, ... N=Dec in Annual Overview
    # Income row = 5 in Annual Overview, Expense total row = 19
    ws_dash.cell(row=r, column=3).value = f"='Annual Overview'!{ao_col}5"
    ws_dash.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws_dash.cell(row=r, column=3).font = font_money
    ws_dash.cell(row=r, column=3).fill = bg
    ws_dash.cell(row=r, column=3).border = thin_border
    ws_dash.cell(row=r, column=3).alignment = align_right

    ws_dash.cell(row=r, column=4).value = f"='Annual Overview'!{ao_col}19"
    ws_dash.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws_dash.cell(row=r, column=4).font = font_money
    ws_dash.cell(row=r, column=4).fill = bg
    ws_dash.cell(row=r, column=4).border = thin_border
    ws_dash.cell(row=r, column=4).alignment = align_right

    ws_dash.cell(row=r, column=5).value = f"=C{r}-D{r}"
    ws_dash.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws_dash.cell(row=r, column=5).font = font_money
    ws_dash.cell(row=r, column=5).fill = bg
    ws_dash.cell(row=r, column=5).border = thin_border
    ws_dash.cell(row=r, column=5).alignment = align_right

    ws_dash.cell(row=r, column=6).value = f'=IF(C{r}=0,0,E{r}/C{r})'
    ws_dash.cell(row=r, column=6).number_format = PERCENT_FMT
    ws_dash.cell(row=r, column=6).font = font_money
    ws_dash.cell(row=r, column=6).fill = bg
    ws_dash.cell(row=r, column=6).border = thin_border
    ws_dash.cell(row=r, column=6).alignment = align_center

# Net savings conditional formatting in monthly comparison
net_range = f"E29:E40"
ws_dash.conditional_formatting.add(net_range,
    CellIsRule(operator='greaterThan', formula=['0'], fill=emerald_fill, font=font_positive))
ws_dash.conditional_formatting.add(net_range,
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill_bg, font=font_negative))

# Annual total row
atr = 41
ws_dash.cell(row=atr, column=2, value="ANNUAL TOTAL").font = font_white_bold
ws_dash.cell(row=atr, column=2).fill = navy_fill
ws_dash.cell(row=atr, column=2).border = medium_border
ws_dash.cell(row=atr, column=2).alignment = align_center
for c in range(3, 7):
    ws_dash.cell(row=atr, column=c).fill = navy_fill
    ws_dash.cell(row=atr, column=c).border = medium_border
    ws_dash.cell(row=atr, column=c).font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    ws_dash.cell(row=atr, column=c).alignment = align_center
ws_dash.cell(row=atr, column=3).value = "=SUM(C29:C40)"
ws_dash.cell(row=atr, column=3).number_format = CURRENCY_FMT
ws_dash.cell(row=atr, column=4).value = "=SUM(D29:D40)"
ws_dash.cell(row=atr, column=4).number_format = CURRENCY_FMT
ws_dash.cell(row=atr, column=5).value = "=C41-D41"
ws_dash.cell(row=atr, column=5).number_format = CURRENCY_FMT
ws_dash.cell(row=atr, column=6).value = '=IF(C41=0,0,E41/C41)'
ws_dash.cell(row=atr, column=6).number_format = PERCENT_FMT
for c in [7]:
    ws_dash.cell(row=atr, column=c).fill = navy_fill
    ws_dash.cell(row=atr, column=c).border = medium_border

ws_dash.freeze_panes = 'B3'


# ───────────────────────────────────────────────────────────────
# SHEET 3: MONTHLY BUDGET
# ───────────────────────────────────────────────────────────────
ws_budget = wb.create_sheet("Monthly Budget")
ws_budget.sheet_properties.tabColor = "2563EB"

set_print_settings(ws_budget)

col_widths_budget = {'A': 3, 'B': 24, 'C': 18, 'D': 18, 'E': 18, 'F': 16, 'G': 3}
for col, w in col_widths_budget.items():
    ws_budget.column_dimensions[col].width = w

# Title
ws_budget.merge_cells('A1:G2')
ws_budget['A1'].value = "MONTHLY BUDGET PLANNER"
ws_budget['A1'].font = Font(name='Calibri', size=22, bold=True, color=WHITE)
ws_budget['A1'].fill = navy_fill
ws_budget['A1'].alignment = align_center
for r in range(1, 3):
    for c in range(1, 8):
        ws_budget.cell(row=r, column=c).fill = navy_fill

# Month selector info
ws_budget.merge_cells('B3:F3')
ws_budget.cell(row=3, column=2).value = "Enter your budgeted amounts in column C. Actual spending is calculated automatically from the Transactions sheet."
ws_budget.cell(row=3, column=2).font = Font(name='Calibri', size=10, italic=True, color=GRAY_DARK)
ws_budget.cell(row=3, column=2).alignment = align_center
ws_budget.row_dimensions[3].height = 28

# Headers row 4
budget_headers = ["Category", "Budgeted Amount", "Actual Spent", "Difference", "Status"]
row = 4
for i, h in enumerate(budget_headers):
    cell = ws_budget.cell(row=row, column=i + 2)
    cell.value = h
    cell.font = font_header
    cell.fill = navy_fill
    cell.alignment = align_center
    cell.border = header_border

# Category rows
for idx, cat in enumerate(CATEGORIES):
    r = 5 + idx
    bg = light_fill if idx % 2 == 0 else white_fill
    ws_budget.row_dimensions[r].height = 28

    # Category name
    ws_budget.cell(row=r, column=2, value=cat)
    ws_budget.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=NAVY)
    ws_budget.cell(row=r, column=2).fill = bg
    ws_budget.cell(row=r, column=2).border = thin_border
    ws_budget.cell(row=r, column=2).alignment = align_left

    # Budgeted Amount (user input) - pre-fill with sample
    ws_budget.cell(row=r, column=3, value=SAMPLE_BUDGETS[cat])
    ws_budget.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws_budget.cell(row=r, column=3).font = font_money
    ws_budget.cell(row=r, column=3).fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type='solid') if True else bg  # light yellow to indicate editable
    ws_budget.cell(row=r, column=3).border = Border(
        left=Side(style='thin', color=AMBER),
        right=Side(style='thin', color=AMBER),
        top=Side(style='thin', color=AMBER),
        bottom=Side(style='thin', color=AMBER),
    )
    ws_budget.cell(row=r, column=3).alignment = align_right

    # Actual Spent (SUMIFS from Transactions where category matches, negative amounts)
    ws_budget.cell(row=r, column=4).value = f'=SUMPRODUCT((Transactions!C2:C501=B{r})*(Transactions!D2:D501<0)*Transactions!D2:D501)*-1'
    ws_budget.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws_budget.cell(row=r, column=4).font = font_money
    ws_budget.cell(row=r, column=4).fill = bg
    ws_budget.cell(row=r, column=4).border = thin_border
    ws_budget.cell(row=r, column=4).alignment = align_right

    # Difference = Budgeted - Actual
    ws_budget.cell(row=r, column=5).value = f'=C{r}-D{r}'
    ws_budget.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws_budget.cell(row=r, column=5).font = font_money
    ws_budget.cell(row=r, column=5).fill = bg
    ws_budget.cell(row=r, column=5).border = thin_border
    ws_budget.cell(row=r, column=5).alignment = align_right

    # Status
    ws_budget.cell(row=r, column=6).value = f'=IF(D{r}=0,"-",IF(D{r}<=C{r}*0.9,"Under Budget",IF(D{r}<=C{r},"Near Limit","OVER BUDGET")))'
    ws_budget.cell(row=r, column=6).font = font_body
    ws_budget.cell(row=r, column=6).fill = bg
    ws_budget.cell(row=r, column=6).border = thin_border
    ws_budget.cell(row=r, column=6).alignment = align_center

# Conditional formatting for Difference column
diff_range_b = "E5:E16"
ws_budget.conditional_formatting.add(diff_range_b,
    CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=emerald_fill, font=font_positive))
ws_budget.conditional_formatting.add(diff_range_b,
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill_bg, font=font_negative))

# Status conditional formatting
status_range_b = "F5:F16"
ws_budget.conditional_formatting.add(status_range_b,
    CellIsRule(operator='equal', formula=['"Under Budget"'], fill=emerald_fill, font=font_positive))
ws_budget.conditional_formatting.add(status_range_b,
    CellIsRule(operator='equal', formula=['"OVER BUDGET"'], fill=red_fill_bg, font=font_negative))
ws_budget.conditional_formatting.add(status_range_b,
    CellIsRule(operator='equal', formula=['"Near Limit"'], fill=amber_fill,
               font=Font(name='Calibri', size=11, bold=True, color=AMBER)))

# Totals row
tr = 17
ws_budget.cell(row=tr, column=2, value="TOTAL").font = font_white_bold
ws_budget.cell(row=tr, column=2).fill = navy_fill
ws_budget.cell(row=tr, column=2).border = medium_border
ws_budget.cell(row=tr, column=2).alignment = align_center
for c in range(3, 7):
    ws_budget.cell(row=tr, column=c).fill = navy_fill
    ws_budget.cell(row=tr, column=c).border = medium_border
    ws_budget.cell(row=tr, column=c).font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    ws_budget.cell(row=tr, column=c).alignment = align_center

ws_budget.cell(row=tr, column=3).value = "=SUM(C5:C16)"
ws_budget.cell(row=tr, column=3).number_format = CURRENCY_FMT
ws_budget.cell(row=tr, column=4).value = "=SUM(D5:D16)"
ws_budget.cell(row=tr, column=4).number_format = CURRENCY_FMT
ws_budget.cell(row=tr, column=5).value = "=SUM(E5:E16)"
ws_budget.cell(row=tr, column=5).number_format = CURRENCY_FMT

# Data validation for categories (users might want to rename)
# Category dropdown for if they want to reorder (optional - categories are fixed labels here)

ws_budget.freeze_panes = 'B5'


# ───────────────────────────────────────────────────────────────
# SHEET 4: TRANSACTIONS
# ───────────────────────────────────────────────────────────────
ws_trans = wb.create_sheet("Transactions")
ws_trans.sheet_properties.tabColor = "F59E0B"

set_print_settings(ws_trans)

col_widths_trans = {'A': 14, 'B': 30, 'C': 20, 'D': 14, 'E': 18, 'F': 25}
for col, w in col_widths_trans.items():
    ws_trans.column_dimensions[col].width = w

# Title - separate from data for cleaner formulas
# Actually, let's keep headers at row 1 so formulas referencing D2:D501 work cleanly

# Headers
trans_headers = ["Date", "Description", "Category", "Amount", "Payment Method", "Notes"]
for i, h in enumerate(trans_headers):
    cell = ws_trans.cell(row=1, column=i + 1)
    cell.value = h
    cell.font = font_header
    cell.fill = navy_fill
    cell.alignment = align_center
    cell.border = header_border

ws_trans.row_dimensions[1].height = 32

# Pre-format 500 data rows
for r in range(2, 502):
    bg = light_fill if r % 2 == 0 else white_fill
    for c in range(1, 7):
        cell = ws_trans.cell(row=r, column=c)
        cell.fill = bg
        cell.border = thin_border
        cell.font = font_body

    # Date column formatting
    ws_trans.cell(row=r, column=1).number_format = DATE_FMT
    ws_trans.cell(row=r, column=1).alignment = align_center

    # Amount formatting
    ws_trans.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws_trans.cell(row=r, column=4).alignment = align_right

    # Center-align category and payment method
    ws_trans.cell(row=r, column=3).alignment = align_center
    ws_trans.cell(row=r, column=5).alignment = align_center

# Data Validation: Category dropdown
cat_list = '"' + ','.join(["Income"] + CATEGORIES) + '"'
dv_category = DataValidation(type="list", formula1=cat_list, allow_blank=True)
dv_category.error = "Please select a valid category from the dropdown."
dv_category.errorTitle = "Invalid Category"
dv_category.prompt = "Select a category"
dv_category.promptTitle = "Category"
ws_trans.add_data_validation(dv_category)
dv_category.add('C2:C501')

# Data Validation: Payment Method dropdown
pm_list = '"' + ','.join(PAYMENT_METHODS) + '"'
dv_payment = DataValidation(type="list", formula1=pm_list, allow_blank=True)
dv_payment.error = "Please select a valid payment method."
dv_payment.errorTitle = "Invalid Payment Method"
dv_payment.prompt = "Select payment method"
dv_payment.promptTitle = "Payment Method"
ws_trans.add_data_validation(dv_payment)
dv_payment.add('E2:E501')

# Date validation
dv_date = DataValidation(type="date", allow_blank=True)
dv_date.error = "Please enter a valid date (MM/DD/YYYY)."
dv_date.errorTitle = "Invalid Date"
ws_trans.add_data_validation(dv_date)
dv_date.add('A2:A501')

# Add sample transactions to show how it works
sample_transactions = [
    (datetime.date(2026, 2, 1), "Monthly Salary", "Income", 5500.00, "Bank Transfer", "Net pay after taxes"),
    (datetime.date(2026, 2, 1), "Rent Payment", "Housing", -1500.00, "Bank Transfer", "February rent"),
    (datetime.date(2026, 2, 2), "Grocery Store", "Food & Groceries", -127.43, "Debit Card", "Weekly groceries"),
    (datetime.date(2026, 2, 3), "Electric Bill", "Utilities", -95.20, "Bank Transfer", "February electric"),
    (datetime.date(2026, 2, 3), "Gas Station", "Transportation", -48.50, "Credit Card", "Full tank"),
    (datetime.date(2026, 2, 5), "Netflix Subscription", "Entertainment", -15.99, "Credit Card", "Monthly streaming"),
    (datetime.date(2026, 2, 5), "Car Insurance", "Insurance", -145.00, "Bank Transfer", "Monthly premium"),
    (datetime.date(2026, 2, 7), "Pharmacy", "Healthcare", -32.50, "Debit Card", "Prescriptions"),
    (datetime.date(2026, 2, 8), "Student Loan", "Debt Payments", -350.00, "Bank Transfer", "Monthly payment"),
    (datetime.date(2026, 2, 10), "Haircut", "Personal", -35.00, "Cash", ""),
    (datetime.date(2026, 2, 12), "Online Course", "Education", -49.99, "Credit Card", "Python masterclass"),
    (datetime.date(2026, 2, 14), "Valentine's Dinner", "Entertainment", -85.00, "Credit Card", "Restaurant"),
    (datetime.date(2026, 2, 15), "Freelance Payment", "Income", 800.00, "Bank Transfer", "Logo design project"),
    (datetime.date(2026, 2, 15), "Savings Transfer", "Savings", -500.00, "Bank Transfer", "Monthly savings"),
    (datetime.date(2026, 2, 18), "Water Bill", "Utilities", -45.00, "Bank Transfer", "February water"),
    (datetime.date(2026, 2, 20), "Grocery Store", "Food & Groceries", -98.75, "Debit Card", "Weekly groceries"),
]

for i, (date, desc, cat, amt, method, notes) in enumerate(sample_transactions):
    r = 2 + i
    ws_trans.cell(row=r, column=1, value=date)
    ws_trans.cell(row=r, column=2, value=desc)
    ws_trans.cell(row=r, column=3, value=cat)
    ws_trans.cell(row=r, column=4, value=amt)
    ws_trans.cell(row=r, column=5, value=method)
    ws_trans.cell(row=r, column=6, value=notes)

# Conditional formatting: negative amounts in red, positive in green
ws_trans.conditional_formatting.add('D2:D501',
    CellIsRule(operator='greaterThan', formula=['0'], font=font_positive))
ws_trans.conditional_formatting.add('D2:D501',
    CellIsRule(operator='lessThan', formula=['0'], font=font_negative))

ws_trans.freeze_panes = 'A2'

# Auto-filter
ws_trans.auto_filter.ref = "A1:F501"


# ───────────────────────────────────────────────────────────────
# SHEET 5: ANNUAL OVERVIEW
# ───────────────────────────────────────────────────────────────
ws_annual = wb.create_sheet("Annual Overview")
ws_annual.sheet_properties.tabColor = "8B5CF6"

set_print_settings(ws_annual)

# Column widths
ws_annual.column_dimensions['A'].width = 3
ws_annual.column_dimensions['B'].width = 22
for i, m in enumerate(MONTHS):
    ws_annual.column_dimensions[get_column_letter(i + 3)].width = 14
ws_annual.column_dimensions[get_column_letter(15)].width = 16  # Annual Total col O
ws_annual.column_dimensions[get_column_letter(16)].width = 14  # Monthly Avg col P

# Title
ws_annual.merge_cells('A1:P2')
ws_annual['A1'].value = "ANNUAL FINANCIAL OVERVIEW"
ws_annual['A1'].font = Font(name='Calibri', size=22, bold=True, color=WHITE)
ws_annual['A1'].fill = navy_fill
ws_annual['A1'].alignment = align_center
for r in range(1, 3):
    for c in range(1, 17):
        ws_annual.cell(row=r, column=c).fill = navy_fill

# Year label
ws_annual.merge_cells('B3:P3')
ws_annual.cell(row=3, column=2).value = "Year: 2026  |  Enter monthly totals or let formulas auto-calculate from the Transactions sheet"
ws_annual.cell(row=3, column=2).font = Font(name='Calibri', size=10, italic=True, color=GRAY_DARK)
ws_annual.cell(row=3, column=2).alignment = align_center

# Headers row 4
row = 4
ws_annual.cell(row=row, column=2, value="Category").font = font_header
ws_annual.cell(row=row, column=2).fill = navy_fill
ws_annual.cell(row=row, column=2).alignment = align_center
ws_annual.cell(row=row, column=2).border = header_border

for i, m in enumerate(MONTHS):
    cell = ws_annual.cell(row=row, column=i + 3)
    cell.value = m
    cell.font = font_header
    cell.fill = navy_fill
    cell.alignment = align_center
    cell.border = header_border

ws_annual.cell(row=row, column=15, value="Annual Total").font = font_header
ws_annual.cell(row=row, column=15).fill = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type='solid')
ws_annual.cell(row=row, column=15).alignment = align_center
ws_annual.cell(row=row, column=15).border = header_border

ws_annual.cell(row=row, column=16, value="Monthly Avg").font = font_header
ws_annual.cell(row=row, column=16).fill = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type='solid')
ws_annual.cell(row=row, column=16).alignment = align_center
ws_annual.cell(row=row, column=16).border = header_border

# Income row
row = 5
make_section_header(ws_annual, row, 2, 2, "INCOME", fill=PatternFill(start_color=EMERALD, end_color=EMERALD, fill_type='solid'))
ws_annual.cell(row=row, column=2).font = Font(name='Calibri', size=11, bold=True, color=WHITE)

for mi in range(12):
    col = mi + 3
    month_num = mi + 1
    # SUMPRODUCT to get income for specific month from Transactions
    ws_annual.cell(row=row, column=col).value = (
        f'=SUMPRODUCT((MONTH(Transactions!A2:A501)={month_num})'
        f'*(Transactions!C2:C501="Income")'
        f'*Transactions!D2:D501)'
    )
    ws_annual.cell(row=row, column=col).number_format = CURRENCY_FMT
    ws_annual.cell(row=row, column=col).font = font_money
    ws_annual.cell(row=row, column=col).fill = emerald_fill
    ws_annual.cell(row=row, column=col).border = thin_border
    ws_annual.cell(row=row, column=col).alignment = align_right

# Annual total for income
ws_annual.cell(row=row, column=15).value = f"=SUM(C{row}:N{row})"
ws_annual.cell(row=row, column=15).number_format = CURRENCY_FMT
ws_annual.cell(row=row, column=15).font = font_money_bold
ws_annual.cell(row=row, column=15).fill = emerald_fill
ws_annual.cell(row=row, column=15).border = thin_border
ws_annual.cell(row=row, column=15).alignment = align_right

# Monthly average
ws_annual.cell(row=row, column=16).value = f'=IF(COUNTIF(C{row}:N{row},"<>0")=0,0,O{row}/COUNTIF(C{row}:N{row},"<>0"))'
ws_annual.cell(row=row, column=16).number_format = CURRENCY_FMT
ws_annual.cell(row=row, column=16).font = font_money
ws_annual.cell(row=row, column=16).fill = emerald_fill
ws_annual.cell(row=row, column=16).border = thin_border
ws_annual.cell(row=row, column=16).alignment = align_right

# Expense categories (rows 6-17, but let's add a small "Expenses" section header)
row = 6
# Section header for expenses
ws_annual.cell(row=row, column=2, value="EXPENSES").font = Font(name='Calibri', size=10, bold=True, color=WHITE)
ws_annual.cell(row=row, column=2).fill = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
ws_annual.cell(row=row, column=2).alignment = align_center
ws_annual.cell(row=row, column=2).border = thin_border
for c in range(3, 17):
    ws_annual.cell(row=row, column=c).fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type='solid')
    ws_annual.cell(row=row, column=c).border = thin_border

for idx, cat in enumerate(CATEGORIES):
    r = 7 + idx
    bg = light_fill if idx % 2 == 0 else white_fill
    ws_annual.cell(row=r, column=2, value=cat)
    ws_annual.cell(row=r, column=2).font = Font(name='Calibri', size=11, color=NAVY)
    ws_annual.cell(row=r, column=2).fill = bg
    ws_annual.cell(row=r, column=2).border = thin_border
    ws_annual.cell(row=r, column=2).alignment = align_left

    for mi in range(12):
        col = mi + 3
        month_num = mi + 1
        ws_annual.cell(row=r, column=col).value = (
            f'=SUMPRODUCT((MONTH(Transactions!A2:A501)={month_num})'
            f'*(Transactions!C2:C501="{cat}")'
            f'*(Transactions!D2:D501<0)*Transactions!D2:D501)*-1'
        )
        ws_annual.cell(row=r, column=col).number_format = CURRENCY_FMT
        ws_annual.cell(row=r, column=col).font = font_money
        ws_annual.cell(row=r, column=col).fill = bg
        ws_annual.cell(row=r, column=col).border = thin_border
        ws_annual.cell(row=r, column=col).alignment = align_right

    # Annual total
    ws_annual.cell(row=r, column=15).value = f"=SUM(C{r}:N{r})"
    ws_annual.cell(row=r, column=15).number_format = CURRENCY_FMT
    ws_annual.cell(row=r, column=15).font = font_money_bold
    ws_annual.cell(row=r, column=15).fill = bg
    ws_annual.cell(row=r, column=15).border = thin_border
    ws_annual.cell(row=r, column=15).alignment = align_right

    # Monthly average
    ws_annual.cell(row=r, column=16).value = f'=IF(COUNTIF(C{r}:N{r},"<>0")=0,0,O{r}/COUNTIF(C{r}:N{r},"<>0"))'
    ws_annual.cell(row=r, column=16).number_format = CURRENCY_FMT
    ws_annual.cell(row=r, column=16).font = font_money
    ws_annual.cell(row=r, column=16).fill = bg
    ws_annual.cell(row=r, column=16).border = thin_border
    ws_annual.cell(row=r, column=16).alignment = align_right

# Total Expenses row
ter = 19
ws_annual.cell(row=ter, column=2, value="TOTAL EXPENSES").font = font_white_bold
ws_annual.cell(row=ter, column=2).fill = navy_fill
ws_annual.cell(row=ter, column=2).border = medium_border
ws_annual.cell(row=ter, column=2).alignment = align_center
for col in range(3, 17):
    ws_annual.cell(row=ter, column=col).fill = navy_fill
    ws_annual.cell(row=ter, column=col).border = medium_border
    ws_annual.cell(row=ter, column=col).font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    ws_annual.cell(row=ter, column=col).alignment = align_right
    if col <= 14:
        ws_annual.cell(row=ter, column=col).value = f"=SUM({get_column_letter(col)}7:{get_column_letter(col)}18)"
    elif col == 15:
        ws_annual.cell(row=ter, column=col).value = f"=SUM(C{ter}:N{ter})"
    elif col == 16:
        ws_annual.cell(row=ter, column=col).value = f'=IF(COUNTIF(C{ter}:N{ter},"<>0")=0,0,O{ter}/COUNTIF(C{ter}:N{ter},"<>0"))'
    ws_annual.cell(row=ter, column=col).number_format = CURRENCY_FMT

# Net Savings row
nsr = 20
ws_annual.merge_cells(f'B{nsr}:B{nsr}')
ws_annual.cell(row=nsr, column=2, value="NET SAVINGS").font = Font(name='Calibri', size=12, bold=True, color=WHITE)
ws_annual.cell(row=nsr, column=2).fill = PatternFill(start_color=EMERALD, end_color=EMERALD, fill_type='solid')
ws_annual.cell(row=nsr, column=2).border = medium_border
ws_annual.cell(row=nsr, column=2).alignment = align_center
for col in range(3, 17):
    ws_annual.cell(row=nsr, column=col).fill = PatternFill(start_color=EMERALD, end_color=EMERALD, fill_type='solid')
    ws_annual.cell(row=nsr, column=col).border = medium_border
    ws_annual.cell(row=nsr, column=col).font = Font(name='Calibri', size=11, bold=True, color=WHITE)
    ws_annual.cell(row=nsr, column=col).alignment = align_right
    cl = get_column_letter(col)
    ws_annual.cell(row=nsr, column=col).value = f"={cl}5-{cl}19"
    ws_annual.cell(row=nsr, column=col).number_format = CURRENCY_FMT

# Savings Rate row
srr = 21
ws_annual.cell(row=srr, column=2, value="SAVINGS RATE").font = Font(name='Calibri', size=11, bold=True, color=NAVY)
ws_annual.cell(row=srr, column=2).fill = light_fill
ws_annual.cell(row=srr, column=2).border = thin_border
ws_annual.cell(row=srr, column=2).alignment = align_center
for col in range(3, 17):
    cl = get_column_letter(col)
    ws_annual.cell(row=srr, column=col).value = f'=IF({cl}5=0,0,{cl}20/{cl}5)'
    ws_annual.cell(row=srr, column=col).number_format = PERCENT_FMT
    ws_annual.cell(row=srr, column=col).font = Font(name='Calibri', size=11, bold=True, color=NAVY)
    ws_annual.cell(row=srr, column=col).fill = light_fill
    ws_annual.cell(row=srr, column=col).border = thin_border
    ws_annual.cell(row=srr, column=col).alignment = align_center

# Conditional formatting on net savings row
for col in range(3, 17):
    cl = get_column_letter(col)
    cell_ref = f"{cl}{nsr}"
    ws_annual.conditional_formatting.add(cell_ref,
        CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                   fill=PatternFill(start_color=EMERALD, end_color=EMERALD, fill_type='solid'),
                   font=Font(name='Calibri', size=11, bold=True, color=WHITE)))
    ws_annual.conditional_formatting.add(cell_ref,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(start_color=RED, end_color=RED, fill_type='solid'),
                   font=Font(name='Calibri', size=11, bold=True, color=WHITE)))

ws_annual.freeze_panes = 'C5'


# ═══════════════════════════════════════════════════════════════
# FINAL TOUCHES
# ═══════════════════════════════════════════════════════════════

# Set the Instructions sheet as the active sheet when opened
wb.active = 0

# Save
output_path = "/Users/timdunn/mobile_app_ideas/etsy-templates/monthly-budget-tracker.xlsx"
wb.save(output_path)
print(f"SUCCESS: Workbook saved to {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Transactions sheet has {len(sample_transactions)} sample rows pre-filled")
print(f"Categories: {len(CATEGORIES)}")
print(f"Pre-formatted transaction rows: 500")

